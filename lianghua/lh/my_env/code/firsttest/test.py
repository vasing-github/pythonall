from enum import verify

import openpyxl
import requests
from openpyxl.styles import Font
import time
import os

# 配置参数
INPUT_FILE = "input1.xlsx"  # 输入文件路径
OUTPUT_FILE = "output.xlsx"  # 输出文件路径
API_URL = "https://a.xinzhi.space/h5/Map/suggestion"


headers = {
    # 保持与原始请求完全相同的headers配置
    "Host": "a.xinzhi.space",
    "Connection": "keep-alive",
    "sec-ch-ua-platform": "\"Windows\"",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36 Edg/134.0.0.0",
    "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Microsoft Edge\";v=\"134\"",
    "Content-Type": "application/json",
    "sec-ch-ua-mobile": "?0",
    "ta": "%7B%22#brower%22:%22Microsoft%20Edge%22,%22#browser_version%22:%22134.0.0.0%22,%22#screen_height%22:1044,%22#screen_width%22:2000,%22#os%22:%22Windows%22,%22#ua%22:%22Mozilla/5.0%20(Windows%20NT%2010.0;%20Win64;%20x64)%20AppleWebKit/537.36%20(KHTML,%20like%20Gecko)%20Chrome/134.0.0.0%20Safari/537.36%20Edg/134.0.0.0%22%7D",
    "Accept": "*/*",
    "Origin": "https://xinzhi.space",
    "Sec-Fetch-Site": "same-site",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Dest": "empty",
    "Referer": "https://xinzhi.space/",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6"
}

def init_workbook():
    """初始化或加载已有输出文件"""
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        print("检测到已有输出文件，开启断点续传模式")
    else:
        # 创建新文件并复制输入数据
        wb = openpyxl.Workbook()
        ws = wb.active

        # 复制输入文件数据
        input_wb = openpyxl.load_workbook(INPUT_FILE)
        input_ws = input_wb.active
        for row in input_ws.iter_rows():
            ws.append([cell.value for cell in row])

        # 添加结果列标题
        titles = ["经度", "纬度", "详细地址", "行政区编码", "省份", "城市", "区县"]
        for col, title in enumerate(titles, start=2):
            ws.cell(row=1, column=col, value=title).font = Font(bold=True)
        wb.save(OUTPUT_FILE)
        print("新建输出文件并初始化完成")

    return wb, ws


# 初始化工作簿
wb, ws = init_workbook()

# 遍历处理（从第二行开始）
for row in range(2, ws.max_row + 1):
    # 检查是否已处理：如果经度列（B列）有值则跳过
    if ws.cell(row=row, column=2).value is not None:
        continue

    query = ws.cell(row=row, column=1).value
    if not query:
        continue

    # 构造请求参数（与之前相同）
    params = {
        "channel": "web",
        "client_type": "3",
        "timestamp": "1740015449",
        "version": "1",
        "sign": "89B000F3A1912C7AB80EA4149BE57F75"
    }

    payload = {
        "device": "56c4f61c-04e6-4617-9267-33596bc0991d",
        "query": query,
        "region": "全国"
    }

    try:
        # 发送请求
        response = requests.post(
            API_URL,
            params=params,
            headers=headers,
            json=payload,
            timeout=10,
            verify=False
        )

        # 处理响应
        if response.status_code == 200:
            data = response.json()
            if data['status'] == 0 and data['result']:
                result = data['result'][0]
                # 写入数据到各列
                ws.cell(row=row, column=2, value=result['location']['lng'])
                ws.cell(row=row, column=3, value=result['location']['lat'])
                ws.cell(row=row, column=4, value=result['address'])
                ws.cell(row=row, column=5, value=result['adcode'])
                ws.cell(row=row, column=6, value=result['province'])
                ws.cell(row=row, column=7, value=result['city'])
                ws.cell(row=row, column=8, value=result.get('district', ''))

                print(f"✅ 成功处理：{query}")
            else:
                # 标记错误信息
                ws.cell(row=row, column=2, value=f"错误：{data.get('msg', '未知错误')}")
                print(f"❌ 接口返回错误：{query}")
        else:
            ws.cell(row=row, column=2, value=f"HTTP错误：{response.status_code}")
            print(f"❌ 请求失败：{query}，状态码：{response.status_code}")

    except Exception as e:
        # 记录异常信息
        ws.cell(row=row, column=2, value=f"异常：{str(e)}")
        print(f"⚠️ 处理异常：{query}，错误：{str(e)}")

    # 立即保存进度（每行处理完都保存）
    try:
        wb.save(OUTPUT_FILE)
    except Exception as save_error:
        print(f"💥 文件保存失败：{str(save_error)}，建议检查文件是否被其他程序占用")
        exit(1)

    # 延迟防止高频请求
    time.sleep(1)

print("🎉 全部处理完成！")