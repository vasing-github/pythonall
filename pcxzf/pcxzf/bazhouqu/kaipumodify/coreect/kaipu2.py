# -*- coding: utf-8 -*-
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta
from io import BytesIO
import one_argument_article
import hudongcontent
import requests
from bs4 import BeautifulSoup
from collections import defaultdict
import two_argument_article
current_dir = os.path.dirname(os.path.abspath(__file__))
# 获取项目的根目录
project_root = os.path.abspath(os.path.join(current_dir, os.pardir, os.pardir))
# 将项目根目录添加到 sys.path
sys.path.append(project_root)
from bazhouqu.kaipumodify.cfg import dealtext
from bazhouqu.kaipumodify.cfg import conf
import bazhouqu.kaipumodify.cfg.text as text
from openpyxl import Workbook
from openpyxl import load_workbook
import getcontent2, getcontent
from bazhouqu.kaipumodify.modifyfile import upfile2



token = text.token

bz_gov_id = text.bz_gov_id
jid = text.jid

endDate = datetime.now()
startDate = endDate - timedelta(days=30)
endDate_str = endDate.strftime('%Y-%m-%d')
startDate_str = startDate.strftime('%Y-%m-%d')


def get_cuomin_list():
    cookies = {
        'HWWAFSESID': '49d6373a41e52c2bd8',
        'HWWAFSESTIME': '1716168884097',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=49d6373a41e52c2bd8; HWWAFSESTIME=1716168884097',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0',
        'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Microsoft Edge";v="126"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 2000,
            'current': 1,
        },
        'check': 1,
        'dataState': True,
        'dateType': 1,
        'codeType': 1,
        'custCode': conf.kaipu_area,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'rectifyStatus': 0,
        'searchBox': '',
        'questionLevel': '',
        'labelIds': [],
        'resultType': '',
        'pageType': '',
        'recommendUpdateList': [],
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/listPageByDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    # print(response.text)
    return response.json()['data']['records']


def get_secrit():
    cookies = {
        'Path': '/',
        'Path': '/',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'Path=/; Path=/',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Chromium";v="130", "Microsoft Edge";v="130", "Not?A_Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 100,
            'current': 1,
        },
        'check': 1,
        'resultTypeList': [
            1,
            2,
        ],
        'dateType': 1,
        'codeType': 1,
        'codeTypeVal': [
            conf.kaipu_area,
        ],
        'startDate': startDate_str,
        'endDate': endDate_str,
        'pageType': '',
        'reviewType': 0,
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteInfoLeakageMaster/listByDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )

    return response.json()['data']['records']


def get_link_use_num(token):
    cookies = {
        'HWWAFSESID': 'b092d732d20af40cfc',
        'HWWAFSESTIME': '1709105903507',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=b092d732d20af40cfc; HWWAFSESTIME=1709105903507',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
        'sec-ch-ua': '"Microsoft Edge";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'startDate': startDate_str,
        'endDate': endDate_str,
        'status': 1,
        'dateType': 0,
        'isSure': 0,
        'urlType': '',
        'checkType': '',
        'siteInfo': '',
        'siteCustCode': conf.kaipu_area,
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteDeadLink/getStatisticsByQueryDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )

    a = response.json()['data']['channelNotRectified']
    b = response.json()['data']['homeNotRectified']
    print("link_use_num:", a + b)
    return a + b


def get_out_link(token):
    cookies = {
        'HWWAFSESID': 'b092d732d20af40cfc',
        'HWWAFSESTIME': '1709105903507',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=b092d732d20af40cfc; HWWAFSESTIME=1709105903507',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
        'sec-ch-ua': '"Microsoft Edge";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 100,
            'current': 1,
        },
        'check': 1,
        'dataState': True,
        'dateType': 4,
        'inspectionType': 1,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'codeType': 1,
        'codeTypeVal': [
            conf.kaipu_area,
        ],
        'pageType': '',
        'reviewType': '',
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteOutLink/reviewDetailStatistics',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print("out_link_num:", response.json()['data']['detail']['0'])
    return response.json()['data']['detail']['0']


def get_yinsi_num(token):
    cookies = {
        'HWWAFSESID': 'b092d732d20af40cfc',
        'HWWAFSESTIME': '1709105903507',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=b092d732d20af40cfc; HWWAFSESTIME=1709105903507',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
        'sec-ch-ua': '"Microsoft Edge";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'check': 1,
        'codeType': 1,
        'codeTypeVal': [
            conf.kaipu_area,
        ],
        'resultTypeList': [
            1,
            2,
        ],
        'dateType': 1,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'page': {
            'size': 100,
            'current': 1,
        },
        'pageType': '',
        'reviewType': 0,
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteInfoLeakageMaster/reviewStatistics',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print("yinsi_num:", response.json()['data']['allCount'])
    return response.json()['data']['allCount']


def start_search():
    print(endDate_str)
    cookies = {
        'HWWAFSESID': 'cba5c76094419b2ace',
        'HWWAFSESTIME': '1737293707742',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=40854762319bc867f7; HWWAFSESTIME=1703639525975',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 10,
            'current': 1,
        },
        'check': 1,
        'codeType': 1,
        'dataState': True,
        'custCode': conf.kaipu_area,
        'dateType': 1,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'rectifyStatus': '',
        'searchBox': '',
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/getRectifiedNum',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    # print(response.json())
    if response.status_code != 200:
        dealtoken()
        return start_search()
    waitrefix = response.json()['data']['notRectifiedNum']
    print('waitrefix:', waitrefix)

    return waitrefix


def dealtoken():
    global token
    token = dealtext.deal_kaipu_token()


def send_msg(cuomin, yinsi_num, outnum, kai_link_use_num, uptime=None):
    # 定义要发送的消息内容
    # 获取当前时间，精确到分钟
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" 检测到开普云有新错敏<font color=\"warning\">{cuomin}</font>条\n\n隐私泄露<font color=\"warning\">{yinsi_num}</font>条\n\n外链暗链<font color=\"warning\">{outnum}</font>条\n\n链接不可用<font color=\"warning\">{kai_link_use_num}</font>条\n\n\n开普检测时间：{uptime}\n\n机器人检测时间：{current_time}\n<@WuXiaoLong>\n"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def get_kaipu_uptime(token):
    cookies = {
        'HWWAFSESID': '40854762319bc867f7',
        'HWWAFSESTIME': '1703639525975',
    }
    endDate = datetime.now()
    startDate = endDate - timedelta(days=30)

    # 格式化日期
    endDate_str = endDate.strftime('%Y-%m-%d')
    startDate_str = startDate.strftime('%Y-%m-%d')
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=40854762319bc867f7; HWWAFSESTIME=1703639525975',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 10,
            'current': 1,
        },
        'check': 1,
        'dataState': True,
        'dateType': 1,
        'codeType': 1,
        'custCode': conf.kaipu_area,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'rectifyStatus': 0,
        'searchBox': '',
        'questionLevel': '',
        'labelIds': [],
        'resultType': '',
        'pageType': '',
        'recommendUpdateList': [],
        'protectCode': conf.kaipu_area,
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/listPageByDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print(token)
    print(response.json()['data']['records'][0]['updateTime'])
    return response.json()['data']['records'][0]['updateTime']


def kaipucorrect(ids):
    cookies = {
        'HWWAFSESID': '49d6373a41e52c2bd8',
        'HWWAFSESTIME': '1716168884097',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=49d6373a41e52c2bd8; HWWAFSESTIME=1716168884097',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0',
        'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Microsoft Edge";v="126"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'ids': ids,
        'rectifyStatus': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/updateCheckTypeByIds',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print("开普云标记", response.text)


def get_new_bzid_jid():
    global bz_gov_id, jid
    # print("bz_gov_id token press...........")
    # new_bz_gov_id, new_jid = getjsseion.get_new_cookie()
    # modify_cookie(new_bz_gov_id, new_jid)
    # save_data()
    bz_gov_id, jid = dealtext.jiyue_token()


def add_2_excel_kaipu(item):
    __kaipu_cuo_list = []
    for cuomin in item:
        __kaipu_cuo_list.append(cuomin['id'])
    kaipucorrect(__kaipu_cuo_list)

def add_to_correct(correctlist, correctids, cuomin):
    correctids.append(cuomin['id'])
    correctlist.append((cuomin['sensitiveWords'], cuomin['recommendUpdate'], cuomin['snapshotNew'], cuomin['url'],
                        cuomin['parentUrl'], cuomin['pageTypeMeaning'], cuomin['parentTitle']))
    kaipucorrect(correctids)


def cuo_1_argument(numbers, cuomin, item):
    res = one_argument_article.getcontent(numbers[0], bz_gov_id, jid)

    if res['status'] == -9:
        get_new_bzid_jid()
        res = one_argument_article.getcontent(numbers[0], bz_gov_id, jid)
    res_save = one_argument_article.savearticnews(res, item, bz_gov_id, jid)
    print(res_save)
    if res_save['status'] == 0:
        send_nopage(cuomin['url'])
    else:
        add_2_excel_kaipu(item)


def cuo_2_aruments(numbers, cuomin, item):

    res = two_argument_article.getcontent(numbers[0], numbers[1], bz_gov_id, jid)
    if res['status'] == -9:
        get_new_bzid_jid()
        res = two_argument_article.getcontent(numbers[0], numbers[1], bz_gov_id, jid)
    res_save = two_argument_article.saveorupdate(res, item,bz_gov_id, jid)

    if res_save['status'] == 0:
        send_nopage(cuomin['url'])
    else:
        add_2_excel_kaipu(item)


def cuo_hudong(cuomin, item):
    is_need_up_toke = hudongcontent.start_kaipu(cuomin, bz_gov_id, jid)
    if is_need_up_toke:
        get_new_bzid_jid()
        hudongcontent.start_kaipu(cuomin, bz_gov_id, jid)
    # add_to_correct(correctlist, correctids, cuomin)
    add_2_excel_kaipu(item)

def extract_numbers(url):
    # 用斜杠截取 URL
    parts = url.split('/')
    numbers = []

    # 遍历每个部分
    for part in parts:
        # 检查部分是否以数字开头并且长度大于4
        match = re.match(r'^\d{5,}', part)
        if match:
            numbers.append(match.group())

    return numbers


def send_secrit_msg(right_type, right_words, url, wrong, title):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f"自动整改网站隐私泄露，请人工核实。\n\n\n[{title}]({url})\n\n\n号码类型：{right_type}\n\n隐私号码：{wrong}\n\n脱敏号码：{right_words}\n\n\n<@WuXiaoLong>\n<@MingFengWangBaoNing>"
        }
    }
    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def send_excel_correct(url):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" PDF,Word附件类错误暂需人工修改。\n\n\n{url}\n\n\n<@WuXiaoLong>\n<@MingFengWangBaoNing>"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    # print(response.text)


def send_excel_modify_success(parent_url, articleTitle, sensitiveWords, recommendUpdate):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" 本次网站改错包含表格附件修改，需人工**重点**核实\n\n\n[{articleTitle}]({parent_url})\n\n\n错敏词：{sensitiveWords}\n\n推荐词：{recommendUpdate}\n\n\n<@WuXiaoLong>\n<@MingFengWangBaoNing>"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)

def make_gongdan_xlsx(cuomin):
    relative_path = os.path.join(current_dir, '..', 'modifyfile', 'sendfile', conf.send_gongdan_xlsx)
    current_time = datetime.now()  # 格式化时间，精确到分钟
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M")
    wb = load_workbook(relative_path)
    ws = wb.active
    row = ws.max_row
    ws.append([formatted_time,cuomin['url'],cuomin['snapshotNew'],cuomin['sensitiveWords'],cuomin['recommendUpdate'],cuomin['articleTitle'],cuomin['parentUrl'],cuomin['parentTitle']])
    wb.save(relative_path)
def remove_protocol(url):
    # 如果URL为空或None，直接返回
    if not url:
        return url
    # 分割协议和后面的部分，取最后一部分（即去除协议）
    return url.split('://', 1)[-1]
def cuo_excel_word(cuomin, item):
    articleTitle = cuomin['articleTitle']
    parent_url = cuomin['parentUrl']
    parentTitle = cuomin['parentTitle']
    sensitiveWords = cuomin['sensitiveWords']
    recommendUpdate = cuomin['recommendUpdate']

    url = find_matching_href(parent_url, articleTitle)

    if url != None and not url.startswith('http'):
        url = conf.jiyuehua_httpstart + url

    print(f'匹配的href: {url}')
    if url is None or remove_protocol(url) != remove_protocol(cuomin.get('url')):  # 父页面中匹配不到附件地址，说明这是缓存的附件，不是页面中真实展示的附件，提交工单删除缓存附件
        make_gongdan_xlsx(cuomin)
        add_2_excel_kaipu(item)
        return
    # 截取最后一个斜杠后的文件名
    filename = url.rsplit('/', 1)[-1]
    path_start = conf.jiyuehua_pathstart
    path_excel = url.split(path_start, 1)[-1]
    filepath = None
    try:
        filepath = upfile2.download_file(url, filename)
    except Exception as e:
        print(f"An error occurred: {e}")

    if filepath == None:
        return

    numbers = extract_numbers(parent_url)
    # 将提取出的数字转换为整数
    numbers = [int(num) for num in numbers]
    # 获取最后一个数字
    last_number = numbers[-1] if numbers else None

    is_doc = upfile2.modify_file(filename, item)
    if is_doc:
        path_excel = path_excel+'x'

    res = upfile2.uploadfile(jid, bz_gov_id, filename, path_excel, parentTitle, articleTitle, last_number)

    if res["status"] == 404:
        get_new_bzid_jid()
        upfile2.uploadfile(jid, bz_gov_id, filename, path_excel, parentTitle, articleTitle, last_number)
    elif res['desc'] == "参数传递有误！":
        return

    elif res["desc"] == "源文件不存在！":
        print("==============================================\n")
        return


    send_excel_modify_success(parent_url, articleTitle, sensitiveWords, recommendUpdate)
    add_2_excel_kaipu(item)


def deal_cuomin_list(list_cuomin):
    # 创建一个默认字典，值为列表
    result_dict = defaultdict(list)

    # 遍历list，将每个元素添加到result_dict中
    for item in list_cuomin:
        url = item['url']
        result_dict[url].append(item)

    # 将默认字典转换为普通字典
    result_dict = dict(result_dict)
    return  result_dict



def dealcuo():
    kaipu_waitrefix = start_search()
    if kaipu_waitrefix != 0:

        list_cuomin = get_cuomin_list()

        result_dic = deal_cuomin_list(list_cuomin)
        for url , item in result_dic.items():

            cuomin = item[0]
            print(cuomin['sensitiveWords'], cuomin['recommendUpdate'])
            print(url)
            if cuomin['pageType'] == "3" and cuomin['column'] != '县长信箱' and cuomin['column'] != '书记信箱' and cuomin['column'] != '互动交流':  # 表示是文章类型

                numbers = extract_numbers(url)
                # 将提取出的数字转换为整数
                numbers = [int(num) for num in numbers]
                # 判断 URL 类型并返回结果
                if len(numbers) == 1:
                    cuo_1_argument(numbers, cuomin, item)
                elif len(numbers) == 2:
                    cuo_2_aruments(numbers, cuomin, item)
                else:
                    print("未知情况")
                    send_nosee()

            elif cuomin['column'] == '互动交流' or cuomin['column'] == '县长信箱' or cuomin['column'] == '书记信箱':
                try:
                    # cuo_hudong(cuomin, item)
                    pass
                except Exception as e:
                    print(f"An error occurred: {e}")

            elif cuomin['pageType'] == "7" or cuomin['pageType'] == "6" or cuomin['pageType'] == "9":  # 表格类错误
                cuo_excel_word(cuomin, item)
            else:
                send_excel_correct(cuomin['url'])
            print("\n")


def cuo_word(cuomin, correctlist, correctids):
    articleTitle = cuomin['articleTitle']
    parent_url = cuomin['parentUrl']
    parentTitle = cuomin['parentTitle']
    sensitiveWords = cuomin['sensitiveWords']
    recommendUpdate = cuomin['recommendUpdate']

    url = find_matching_href(parent_url, articleTitle)

    if url != None and not url.startswith('http') :
        url = 'http://www.scpc.gov.cn' + url

    print(f'匹配的href: {url}')
    if url == None or url != cuomin['url']:  # 父页面中匹配不到附件地址，说明这是缓存的附件，不是页面中真实展示的附件，提交工单删除缓存附件
        make_gongdan_xlsx(cuomin)
        add_to_correct(correctlist, correctids, cuomin)
        return

    # 截取最后一个斜杠后的文件名
    filename = url.rsplit('/', 1)[-1]
    path_start = 'www.scpc.gov.cn'
    path_excel = url.split(path_start, 1)[-1]
    filepath = None
    try:
        filepath = upfile2.download_file(url, filename)
    except Exception as e:
        print(f"An error occurred: {e}")

    if filepath == None:
        return

    numbers = extract_numbers(parent_url)
    # 将提取出的数字转换为整数
    numbers = [int(num) for num in numbers]
    # 获取最后一个数字
    last_number = numbers[-1] if numbers else None

    upfile2.modify_file(filename, sensitiveWords, recommendUpdate)

    res = upfile2.uploadfile(jid, bz_gov_id, filename, path_excel, parentTitle, articleTitle, last_number)
    if res['desc'] == "参数传递有误！":
        return

    elif res["desc"] == "源文件不存在！":
        print("==============================================\n")
        return
    if res["status"] != 1:
        get_new_bzid_jid()
        code = upfile2.uploadfile(jid, bz_gov_id, filename, path_excel, parentTitle, articleTitle, last_number)

    send_excel_modify_success(parent_url, articleTitle, sensitiveWords, recommendUpdate)
    add_to_correct(correctlist, correctids, cuomin)


def send_nosee():
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" 检测到开普云有新类型错敏，机器人无法修改，请人工核实。\n<@WuXiaoLong>\n"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def send_nopage(url):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" 检测到无效静态页，请人工核实是否放入回收站。\n\n\n{url}\n\n\n<@WuXiaoLong>\n<@MingFengWangBaoNing>"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def send_correct_msg(correctlist):
    key = conf.key_cs

    wb = Workbook()

    ws = wb.active

    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格设置标题和表头
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '机器人改错记录  ' + endDate_str
    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font

    ws.append(['错敏词', '修改词', '快照', '地址', '父地址', '错敏类型', '父标题'])
    # 设置新添加的行的字体为加粗
    for row in [ws[row + 1]]:
        for cell in row:
            cell.font = conf.header_font

    # 遍历字典

    for tup in correctlist:
        ws.append(tup)
    ws.column_dimensions['A'].width = 20
    # 设置列A的宽度为50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    # 保存文件

    formatted_date = endDate.strftime('%Y-%m-%d %H_%M_%S')  # 使用下划线代替冒号
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 指定一个子目录名，例如 'saved_files'
    subdirectory = 'coreectlist'

    # 创建子目录的完整路径
    subdirectory_path = os.path.join(script_dir, subdirectory)

    # 如果子目录不存在，则创建它
    if not os.path.exists(subdirectory_path):
        os.makedirs(subdirectory_path)

    # 构建最终的文件路径
    filename = os.path.join(subdirectory_path, conf.correct_name + formatted_date + '.xlsx')
    wb.save(filename)
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key

    # 加载现有的Excel文件
    wb = load_workbook(filename)
    # 获取活动工作表

    # 将工作簿保存到一个字节流中
    output = BytesIO()
    wb.save(output)
    base_filename = os.path.basename(filename)

    # 准备发送文件，确保只使用文件名
    files = {"file": (base_filename, output.getvalue())}
    # 准备发送文件

    response = requests.post(
        'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=' + key + '&type=file',
        files=files)

    print(response.text)

    # 检查是否有'media_id'在响应中
    if 'media_id' in response.json():
        # 发送消息
        headers = {"Content-Type": "application/json"}
        message = {
            "msgtype": "file",
            "file": {
                "media_id": response.json()["media_id"]
            }
        }
        response = requests.post(url, headers=headers, json=message)

        # 输出响应结果
        print(response.text)
    else:
        print("No media_id in response")


def find_matching_href(url, target_text):
    # 发送HTTP请求获取网页内容
    response = requests.get(url)
    if response.status_code == 404:
        return
    response.raise_for_status()  # 确保请求成功

    # 指定编码规则
    response.encoding = response.apparent_encoding

    # 使用BeautifulSoup解析网页内容
    soup = BeautifulSoup(response.text, 'html.parser')

    # 查找所有的a标签
    a_tags = soup.find_all('a')

    # 遍历所有a标签，查找与给定文本匹配的标签
    for a_tag in a_tags:
        if a_tag.text.replace(" ", "") == target_text.replace(" ", ""):
            return a_tag.get('href')

    return None


def tag_secrit_kaipu(sid):
    cookies = {
        'Path': '/',
        'Path': '/',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'Path=/; Path=/',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Chromium";v="130", "Microsoft Edge";v="130", "Not?A_Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'idList': [
            sid,
        ],
        'reviewType': 2,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteInfoLeakageMaster/updateReviewType',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )

    print(response.text)


def change_secrit_right(s):
    # 身份证号码的正则表达式（假设是中国的身份证号码）
    id_card_pattern = r'^\d{15}$|^\d{17}[\dXx]$'

    # 银行卡号码的正则表达式（假设是16到19位数字）
    bank_card_pattern = r'^\d{16,19}$'

    # 手机号码的正则表达式（假设是中国的手机号）
    phone_number_pattern = r'^1[3-9]\d{9}$'

    if re.match(id_card_pattern, s):
        # 身份证中间8位替换为*号
        return "身份证", s[:6] + '*' * 8 + s[-4:]
    elif re.match(bank_card_pattern, s):
        # 银行卡中间4位替换为*号
        return "银行卡", s[:6] + '*' * 4 + s[-6:]
    elif re.match(phone_number_pattern, s):
        # 手机号码中间4位替换为*号
        return "手机号", s[:3] + '*' * 4 + s[-4:]
    else:
        return send_add_code()


def send_add_code():
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f"自动整改网站隐私泄露.\n\n\n发现新的隐私泄露类型或文章位置或未知错误，需要登录开普云看数据、加代码\n\n\n<@WuXiaoLong>"
        }
    }
    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def deal_secrit_artic(se, wrong, right_words):
    numbers = extract_numbers(se['url'])
    # 将提取出的数字转换为整数
    numbers = [int(num) for num in numbers]
    # 判断 URL 类型并返回结果
    if len(numbers) == 1:
        res = getcontent2.getcontent(numbers[0], bz_gov_id, jid)
        res_save = getcontent2.savearticnews(res, wrong, right_words, bz_gov_id, jid)
    elif len(numbers) == 2:
        res = getcontent.getcontent(numbers[0], numbers[1], bz_gov_id, jid)
        res_save = getcontent.saveorupdate(res, wrong, right_words, bz_gov_id, jid)
    else:
        print("未知情况")
        send_add_code()


def deal_secrit_hudong(se, wrong, right_words):
    res = hudongcontent.secrit_deal(wrong, right_words, se, bz_gov_id, jid)
    if res:
        get_new_bzid_jid()
        hudongcontent.secrit_deal(wrong, right_words, se, bz_gov_id, jid)
    # add_to_correct(correctlist, correctids, cuomin)


def deal_secrit():
    secrits = get_secrit()
    for se in secrits:
        print(se['wrongTerms'])
        print(se['url'])
        # 使用正则表达式分隔字符串
        split_strings = re.split(r',', se['wrongTerms'])
        # 遍历分隔后的字符串
        for s in split_strings:
            print(s)
            right_type, right_words = change_secrit_right(s)
            if se['pageType'] == 3:
                deal_secrit_hudong(se, s, right_words)
                deal_secrit_artic(se, s, right_words)
                send_secrit_msg(right_type, right_words, se['url'], s, se['title'])
                tag_secrit_kaipu(se['id'])
            else:
                send_add_code()


if __name__ == '__main__':
    dealcuo()
    deal_secrit()




    # res = getcontent.getcontent('6603801', '10686931', bz_gov_id, jid)
    # if res['status'] == -9:
    #     get_new_bzid_jid()
    #     res = getcontent.getcontent('6603801', '10686931', bz_gov_id, jid)
    # print(res)
    # res_save = getcontent.saveorupdate(res, '下午17:00', '下午5时', bz_gov_id,
    #                                    jid)
    # print(res_save)

    # res = getcontent2.getcontent('13951566', bz_gov_id, jid)
    #
    # if res['status'] == -9:
    #     get_new_bzid_jid()
    #     res = getcontent2.getcontent('13951566', bz_gov_id, jid)
    # print(res)
    # res_save = getcontent2.savearticnews(res,'下午17:00', '下午5时', bz_gov_id, jid)
    # print(res_save)

# 测试表格类改错
#     articleTitle = '部门整体支出绩效目标表.xlsx'
#     parent_url = "http://www.scpc.gov.cn/public/6603501/13965387.html"
#     parentTitle = '大寨镇2023年部门整体支出绩效评价报告'
#     sensitiveWords = '建立建全'
#     recommendUpdate = '建立健全'
#     url = find_matching_href(parent_url, articleTitle)
#     print(f'匹配的href: {url}')
#     if not url.startswith('http'):
#         url = 'http://www.scpc.gov.cn' + url
#     # 截取最后一个斜杠后的文件名
#     filename = url.rsplit('/', 1)[-1]
#     path_start = 'www.scpc.gov.cn'
#     path_excel = url.split(path_start, 1)[-1]
#     upfile.download_file(url, filename)
#     numbers = extract_numbers(parent_url)
#     # 将提取出的数字转换为整数
#     numbers = [int(num) for num in numbers]
#     # 获取最后一个数字
#     last_number = numbers[-1] if numbers else None
#     upfile.modify_file(filename, sensitiveWords, recommendUpdate)
#     code = upfile.uploadfile(jid, bz_gov_id, filename, path_excel, parentTitle, articleTitle, last_number)
#     if code != 200:
#         print("chaoqi")
#         code = upfile.uploadfile(jid, bz_gov_id, filename, path_excel, parentTitle, articleTitle, last_number)



