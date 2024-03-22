# 引入 Beautiful Soup 模块
from bs4 import BeautifulSoup
import requests
import json
import datetime
import config
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
from openpyxl import Workbook
def getAllbumen(prox):
    url = 'http://www.scpc.gov.cn/public/column/6601841?type=4&action=rel'
    data = {}
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    response = requests.get(url, headers=headers, data=data,proxies=prox)

    # 以 Beautiful Soup 解析 HTML 程序码
    soup = BeautifulSoup(response.text, 'html.parser')

    # 提取所有 ul 标签下的所有 li 元素
    li_elements = []
    for ul in soup.find_all('ul', class_='clearfix'):
        li_elements += ul.find_all('li')

    # 将 li 元素转换为字典
    bumenList = {}
    for li in li_elements:
        a_element = li.find('a')
        if a_element is not None:
            title = a_element.get('title')
            href = a_element.get('href')
            bumenList[title] = href

    # 输出字典
    print(bumenList)
    return bumenList
def getMenniu(prox,bumen):
    if not bumen.startswith('http'):
        bumen = 'http://www.scpc.gov.cn' + bumen
    data = {}
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    response = requests.get(bumen, headers=headers, data=data,proxies=prox)
    # 以 Beautiful Soup 解析 HTML 程序码
    soup = BeautifulSoup(response.text, 'html.parser')
    # 提取 id 为 fdzd_gknr 的 div 标签下的所有 a 标签
    a_elements = soup.find('div', {'id': 'fdzd_gknr'}).find_all('a')
    # 输出所有 a 标签的文本内容和链接地址
    menuList={}
    for a in a_elements:
        # print(f'{a.text}: {a["href"]}')
        if a is not None:
            title = a.text
            href = a["href"]
            menuList[title] = href
    return menuList
def getTime(prox,menu):
    if not menu.startswith('http'):
        menu = 'http://www.scpc.gov.cn' + menu

    data = {}
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    response = requests.get(menu, headers=headers, data=data,proxies=prox)

    # 以 Beautiful Soup 解析 HTML 程序码
    soup = BeautifulSoup(response.text, 'html.parser')
    li_elements = soup.find_all('li', {'class': 'rq'})
    # print(li_elements)
    if len(li_elements) ==1:
        return '无'
    second_li_element = li_elements[1]

    return second_li_element.text
def sendMsgError(tittle,content):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f"{tittle}\n\n\n{content}"
        }
    }
    # 定义企业微信机器人的webhook地址
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=425a53f6-696e-46c1-9d4a-fae8940b136f"

    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))
    # 输出响应结果
    print(response.text)
def sendMsg(msg):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=425a53f6-696e-46c1-9d4a-fae8940b136f"

    # 上传文件
    filename = 'msg.txt'
    with open(filename, "w") as f:
        for key, value in msg.items():
            f.write(f"## {key}\n\n")
            for menu, href in value.items():
                f.write(f"- {menu}: {cacul(href)}天未更新\n")
            f.write("\n")

    filename = 'msg.txt'


    with open('./msg.txt', "r") as f:
        files = {"file": ("msg.txt", f.read())}
        response = requests.post('https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=425a53f6-696e-46c1-9d4a-fae8940b136f&type=file', files=files)
    print(response.text)

    # 检查是否有'media_id'在响应中
    if 'media_id' in response.json():
        # 发送消息
        headers = {"Content-Type": "application/json"}
        message = {
            "msgtype": "file",
            "file": {
                "media_id": response.json()["media_id"]
            }
        }
        response = requests.post(url, headers=headers, json=message)

        # 输出响应结果
        print(response.text)
    else:
        print("No media_id in response")


def cacul(href):
    # 这里传进来的其实是字符串2023-11-7
    date = datetime.datetime.strptime(href, '%Y-%m-%d').date()
    delta = datetime.date.today() - date
    return delta.days

def makeMsg(ws,msg):
    # 获取当前的最大行数，并加1
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格设置标题和表头
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '政府工作报告汇总'

    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = config.tittle_font
    ws.append(['单位名称', '标题', '发布时间'])
    # 设置新添加的行的字体为加粗
    for cell in ws[row + 1]:
        cell.font = config.header_font

    # 遍历字典
    now = ''
    for dup in msg:

        if now != dup[0]:
            ws.append([''])

        ws.append(dup)
        now = dup[0]

def  make_no_content_list(ws,no_content_list):
    # 填充该栏目监测的其他问题
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 1
    ws.append(['该栏目检测中发现的其他问题'])
    ws.append(['1.死链监测：以下单位对应栏目无内容，请上传内容或审核发布'])
    ws.append(['单位名称','栏目名称'])
    for r in [row, row + 2]:
        for cell in ws[r]:
            cell.font = conf.header_font

    for dup in no_content_list:
        ws.append([dup[0],dup[1]])
def make_tj_miss_year(ws,tj_miss_year_dic):
    # 填充该栏目监测的其他问题
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 1

    ws.append(['2.统计信息缺失年份：以下单位缺失对应年份统计信息，请补充'])
    ws.append(['单位名称', '缺失年份'])
    for r in [row, row + 1]:
        for cell in ws[r]:
            cell.font = conf.header_font

    for key, value in tj_miss_year_dic.items():
        # 将key和value中的元素合并成一个列表
        row_data = [key] + value
        # 添加到工作表中
        ws.append(row_data)
def makeExcel(list):
    # 加载现有的Excel文件
    wb = Workbook()
    # 保存文件
    wb.save("政府工作报告.xlsx")
    wb = load_workbook("政府工作报告.xlsx")
    # 获取活动工作表
    ws = wb.active

    makeMsg(ws,list)
    # make_no_content_list(ws,no_content_list)
    # make_tj_miss_year(ws,tj_miss_year_dic)

    # 保存文件
    wb.save("政府工作报告.xlsx")

# 这个方法不是直接取网页数据了，是取网页中的js方法，以及参数，用它来请求每一页的数据，如果js为空  说明就只有一页数据，那就拿到这一页的日期返回去，如果不为空，就用gettimebypage方法返回所有日期列表
def getkeywork(prox,menu,time_list):
    if not menu.startswith('http'):
        menu = 'http://www.scpc.gov.cn' + menu

    data = {}
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    response = requests.get(menu, headers=headers, data=data,proxies=prox)

    # 创建BeautifulSoup对象
    soup = BeautifulSoup(response.text, 'html.parser')
    # 查找所有的<script>标签
    scripts = soup.find_all('script')

    # 遍历所有的<script>标签
    for script in scripts:
        # 检查<script>标签的内容是否为None
        if script.string is not None:
            # 检查<script>标签的内容是否包含感兴趣的JavaScript代码
            if 'Ls.pagination' in script.string:
                # 打印<script>标签的内容

                # 使用正则表达式提取url和data
                url_match = re.search(r'url: "(.*?)"', script.string)
                data_match = re.search(r'data : ({.*?})', script.string, re.DOTALL)
                page_count_match = re.search(r'pageCount:(\d+)', script.string)

                if url_match and data_match and page_count_match:
                    # 提取url
                    url = url_match.group(1)

                    # 提取data
                    data_str = data_match.group(1)

                    page_count = page_count_match.group(1)

                    return  getTimeListByPage(time_list,url,data_str,page_count)


    li_elements = soup.find_all('li', {'class': 'rq'})
    if len(li_elements) == 1:
        return '无'
    time_list.extend(li_elements)
    return time_list
def getTimeListByPage(time_list,url,data_str,pagecount):
    # 存储所有页面的数据
    all_pages_data = []
    # 使用正则表达式提取siteId，catId和file
    site_id_match = re.search(r'siteId : "(.*?)"', data_str)
    cat_id_match = re.search(r'catId: "(.*?)"', data_str)
    file_match = re.search(r'file:"(.*?)"', data_str)
    organ_id_match = re.search(r'organId: "(.*?)"', data_str)

    if site_id_match and cat_id_match and file_match and organ_id_match:
        # 提取siteId，catId和file
        site_id = site_id_match.group(1)
        cat_id = cat_id_match.group(1)
        file = file_match.group(1)
        organid = organ_id_match.group(1)
        # print(f"siteId: {site_id}, catId: {cat_id}, file: {file}")
        for i in range(1, int(pagecount) + 1):
            # 请求的URL
            if not url.startswith('http'):
                url = 'http://www.scpc.gov.cn' + url
            # print(url)
            # 请求的参数
            params = {
                "labelName": "publicInfoList",
                "siteId": site_id,
                "pageSize": "18",
                "pageIndex": i,
                "action": "list",
                "isDate": "true",
                "dateFormat": "yyyy-MM-dd",
                "length": "50",
                "organId": organid,
                "isDetail": "",
                "type": "4",
                "catId": cat_id,
                "cId": "",
                "result": "暂无相关信息",
                "file": file
            }

            # 发送请求
            response = requests.get(url, params=params)

            # 检查响应状态码
            if response.status_code == 200:
                # 将页面数据添加到列表中
                all_pages_data.append(response.text)
                soup = BeautifulSoup(response.text, 'html.parser')
                li_elements = soup.find_all('li', {'class': 'rq'})

                time_list.extend(li_elements)
            else:
                print(f"请求失败，状态码：{response.status_code}")

    return time_list
def get_all_tj_time(href,time_list,key,menu,no_content_list):
    upTime = getkeywork(None, href, time_list)
    # print('uptime', upTime)
    # print(len(time_list))
    if upTime == '无':
        no_content_list.append((key, menu))
    return upTime
def deal_tj_year(final_dic):
    new_data = {}

    # 遍历旧字典
    for unit, dates in final_dic.items():
        if unit == '平昌县人民政府办公室':
            continue
        # 初始化年份集合
        years = set()
        # 遍历日期
        for date in dates:
            # 提取年份
            year = date.text.split('-')[0]
            # 添加年份到集合
            years.add(year)
        # 检查是否包含2018-2022年
        missing_years = [str(year) for year in range(2018, 2023) if str(year) not in years]
        # 如果有缺少的年份，添加到新字典
        if missing_years:
            new_data[unit] = missing_years

    # 打印新字典
    print(new_data)
    return new_data
def deal_menu_dic(menu_Dic,key,menu,href,no_content_list,final_dic):
    upTime = getTime(None, href)
    print('uptime', upTime)
    if upTime == '无':
        no_content_list.append((key, menu))
        return
    days = cacul(upTime)
    print(days)
    # 判断时间差是否超过配置表中规定超期时间
    if days >= conf.up_time_conf[menu]:
        menu_Dic[menu] = upTime
    if menu_Dic:
        final_dic[key] = menu_Dic

def deal_zf_report_url(href,danwei,zf_work_list):
    if not href.startswith('http'):
        href = 'http://www.scpc.gov.cn' + href
    res = config.make_request_get(href)
    div_elmen = res[1].find('div', class_='xxgk_nav_con')
    for ul in div_elmen.find_all('ul', class_='clearfix'):
        # 提取日期
        date = ul.find('li', class_='rq').text
        print(f"日期：{date}")

        # 提取标题
        title = ul.find('a', class_='title').text
        print(f"标题：{title}")
        zf_work_list.append((danwei,title,date))

def startMain():
    bumen_Mennu_list = config.bumen_Mennu_list
    final_dic = {}
    no_content_list = []
    tj_final_dic = {}

    zf_work_list = []
    for key, value in bumen_Mennu_list.items():
        print(key)
        menu_Dic = {}
        tj_time_list = []
        for menu, href in value.items():
            print(menu)
            if '政府' not in key:
                continue
            if menu == '政府工作报告':
                deal_zf_report_url(href,key,zf_work_list)

            # deal_menu_dic(menu_Dic,key,menu,href,no_content_list,final_dic)

    print(zf_work_list)
    # zf_work_list = [('平昌县人民政府办公室', '2023年政府工作报告                                            ', '2023-02-10'), ('平昌县人民政府办公室', '2022年政府工作报告                                            ', '2022-03-30'), ('平昌县人民政府办公室', '2013年政府工作报告                                            ', '2021-07-27'), ('平昌县人民政府办公室', '2014年政府工作报告                                            ', '2021-07-27'), ('平昌县人民政府办公室', '2015年政府工作报告                                            ', '2021-07-27'), ('平昌县人民政府办公室', '2016年政府工作报告                                            ', '2021-07-27'), ('平昌县人民政府办公室', '2017年政府工作报告                                            ', '2021-07-27'), ('平昌县人民政府办公室', '2021年政府工作报告                                            ', '2021-06-25'), ('平昌县人民政府办公室', '2020年政府工作报告                                            ', '2020-07-10'), ('平昌县人民政府办公室', '2019年政府工作报告                                            ', '2019-03-06'), ('平昌县人民政府办公室', '2018年政府工作报告                                            ', '2018-03-13'), ('平昌县白衣镇人民政府', '关于2023年上半年工作总结和下半年工作计划                                            ', '2023-07-17'), ('平昌县白衣镇人民政府', '白衣镇2022年政府工作报告                                            ', '2022-06-10'), ('平昌县白衣镇人民政府', '平昌县白衣镇人民政府2021年工作总结和2022年工作计划                                            ', '2022-03-09'), ('平昌县白衣镇人民政府', '白衣镇2019年依法治理工作总结报告                                            ', '2020-01-06'), ('平昌县涵水镇人民政府', '涵水镇2023年政府工作报告                                            ', '2023-08-31'), ('平昌县涵水镇人民政府', '涵水镇2022年政府工作报告                                            ', '2022-08-31'), ('平昌县涵水镇人民政府', '涵水镇2021年政府工作报告                                            ', '2021-10-25'), ('平昌县涵水镇人民政府', '涵水镇2020年政府工作报告                                            ', '2020-07-22'), ('平昌县涵水镇人民政府', '涵水镇2019年政府工作报告                                            ', '2019-07-03'), ('平昌县涵水镇人民政府', '涵水镇2018年政府工作报告                                            ', '2018-06-05'), ('平昌县岳家镇人民政府', '岳家镇2023年政府工作报告                                            ', '2023-04-23'), ('平昌县岳家镇人民政府', '岳家镇2022年政府工作报告                                            ', '2022-04-15'), ('平昌县岳家镇人民政府', '2021年政府工作报告                                            ', '2021-10-25'), ('平昌县岳家镇人民政府', '2020年政府工作报告                                            ', '2020-06-30'), ('平昌县西兴镇人民政府', '西兴镇2022年政府工作报告                                            ', '2022-12-28'), ('平昌县西兴镇人民政府', '西兴镇2021年政府工作报告                                            ', '2021-12-02'), ('平昌县西兴镇人民政府', '西兴镇2020年政府工作报告                                            ', '2020-12-29'), ('平昌县西兴镇人民政府', '2020年巴中市政府工作报告                                            ', '2020-06-04'), ('平昌县西兴镇人民政府', '西兴镇2019年政府工作报告                                            ', '2019-12-09'), ('平昌县西兴镇人民政府', '西兴镇2018年政府工作报告                                            ', '2018-12-07'), ('平昌县西兴镇人民政府', '政府工作报告（文字实录）                                            ', '2018-03-07'), ('平昌县西兴镇人民政府', '西兴镇2017年政府工作报告                                            ', '2017-12-29'), ('平昌县西兴镇人民政府', '西兴镇2016年政府工作报告                                            ', '2016-12-22'), ('平昌县西兴镇人民政府', '西兴镇2015年政府工作报告                                            ', '2015-12-25'), ('平昌县西兴镇人民政府', '西兴镇2014年政府工作报告                                            ', '2014-12-25'), ('平昌县西兴镇人民政府', '西兴镇2013年政府工作报告                                            ', '2013-12-27'), ('平昌县龙岗镇人民政府', '龙岗镇人民政府2023年政府工作报告                                            ', '2023-03-30'), ('平昌县龙岗镇人民政府', '2022年龙岗镇政府工作报告                                            ', '2022-05-07'), ('平昌县龙岗镇人民政府', '平昌县龙岗镇人民政府2021年工作总结和2022年工作计划                                            ', '2022-03-16'), ('平昌县龙岗镇人民政府', '平昌县龙岗镇人民政府2020年工作总结和2021年工作计划                                            ', '2021-06-08'), ('平昌县龙岗镇人民政府', '平昌县龙岗镇人民政府2020年政府工作报告                                            ', '2020-09-11'), ('平昌县龙岗镇人民政府', '平昌县龙岗镇人民政府2019年工作总结和2020年工作计划                                            ', '2020-05-13'), ('平昌县龙岗镇人民政府', '平昌县龙岗镇人民政府2018年工作总结和2019年工作计划                                            ', '2019-03-13'), ('平昌县土垭镇人民政府', '土垭镇2023年政府工作报告                                            ', '2023-04-04'), ('平昌县土垭镇人民政府', '2022年政府工作报告                                            ', '2022-03-31'), ('平昌县土垭镇人民政府', '2021年政府工作报告                                            ', '2021-11-18'), ('平昌县土垭镇人民政府', '2020年政府工作报告                                            ', '2020-09-10'), ('平昌县土垭镇人民政府', '2019年政府工作报告                                            ', '2019-04-09'), ('平昌县土垭镇人民政府', '2018年政府工作报告                                            ', '2018-03-21'), ('平昌县土垭镇人民政府', '2017年政府工作报告                                            ', '2017-03-31'), ('平昌县佛楼镇人民政府', '佛楼镇2023年政府工作报告                                            ', '2023-04-04'), ('平昌县佛楼镇人民政府', '佛楼镇2022年政府工作报告                                            ', '2022-07-27'), ('平昌县佛楼镇人民政府', '佛楼镇2021年政府工作报告                                            ', '2021-10-23'), ('平昌县佛楼镇人民政府', '佛楼镇2020年政府工作报告                                            ', '2020-09-03'), ('平昌县佛楼镇人民政府', '佛楼镇2019年政府工作报告                                            ', '2019-04-10'), ('平昌县佛楼镇人民政府', '佛楼镇人民政府2018年政府工作报告                                            ', '2018-03-21'), ('平昌县佛楼镇人民政府', '佛楼镇人民政府2017年政府工作报告                                            ', '2017-03-30'), ('平昌县佛楼镇人民政府', '佛楼镇2016年政府工作报告                                            ', '2016-10-24'), ('平昌县佛楼镇人民政府', '佛楼镇人民政府2014年政府工作报告                                            ', '2014-03-28'), ('平昌县佛楼镇人民政府', '佛楼镇人民政府2013年政府工作报告                                            ', '2013-03-09'), ('平昌县响滩镇人民政府', '响滩镇人民政府2023年政府工作报告                                            ', '2023-04-21'), ('平昌县响滩镇人民政府', '响滩镇人民政府2022年政府工作报告                                            ', '2022-04-13'), ('平昌县响滩镇人民政府', '响滩镇人民政府2021年政府工作报告                                            ', '2021-10-27'), ('平昌县响滩镇人民政府', '响滩镇人民政府2021年度工作总结                                            ', '2021-09-22'), ('平昌县响滩镇人民政府', '响滩镇人民政府2020年政府工作报告                                            ', '2020-09-22'), ('平昌县响滩镇人民政府', '响滩镇人民政府2019年政府工作报告                                            ', '2019-04-24'), ('平昌县大寨镇人民政府', '大寨镇人民政府2023年政府工作报告                                            ', '2023-01-03'), ('平昌县大寨镇人民政府', '大寨镇2022年政府工作报告                                            ', '2022-07-18'), ('平昌县大寨镇人民政府', '大寨镇2021年政府工作报告                                            ', '2021-12-07'), ('平昌县大寨镇人民政府', '大寨镇2020年政府工作报告                                            ', '2020-09-20'), ('平昌县大寨镇人民政府', '大寨镇2019年政府工作报告                                            ', '2020-05-07'), ('平昌县大寨镇人民政府', '大寨镇2018年政府工作报告                                            ', '2018-03-30'), ('平昌县大寨镇人民政府', '大寨镇2017年政府工作报告                                            ', '2017-03-29'), ('平昌县大寨镇人民政府', '大寨镇2016年政府工作报告                                            ', '2016-04-05'), ('平昌县大寨镇人民政府', '大寨镇2015年政府工作报告                                            ', '2015-05-05'), ('平昌县大寨镇人民政府', '大寨镇2014年政府工作报告                                            ', '2014-03-31'), ('平昌县大寨镇人民政府', '大寨镇2012年政府工作报告                                            ', '2012-06-14'), ('平昌县驷马镇人民政府', '平昌县驷马镇人民政府2023年政府工作报告                                            ', '2023-05-31'), ('平昌县驷马镇人民政府', '驷马镇2022年政府工作报告                                            ', '2022-07-25'), ('平昌县驷马镇人民政府', '驷马镇2021年政府工作报告                                            ', '2021-11-15'), ('平昌县青云镇人民政府', '青云镇人民政府2023年政府工作报告                                            ', '2023-03-30'), ('平昌县青云镇人民政府', '青云镇2022年政府工作报告                                            ', '2022-06-10'), ('平昌县青云镇人民政府', '青云镇2021年政府工作报告                                            ', '2021-12-29'), ('平昌县青云镇人民政府', '青云镇2020年政府工作报告                                            ', '2021-02-26'), ('平昌县青云镇人民政府', '关于2019年政府工作报告的说明                                            ', '2020-05-04'), ('平昌县兰草镇人民政府', '兰草镇人民政府2023年政府工作报告                                            ', '2023-04-18'), ('平昌县兰草镇人民政府', '兰草镇人民政府2022年政府工作报告                                            ', '2022-04-11'), ('平昌县兰草镇人民政府', '兰草镇人民政府2021年政府工作报告                                            ', '2021-10-26'), ('平昌县兰草镇人民政府', '兰草镇人民政府2020年政府工作报告                                            ', '2021-02-01'), ('平昌县兰草镇人民政府', '兰草镇人民政府2019年政府工作报告                                            ', '2019-04-23'), ('平昌县兰草镇人民政府', '兰草镇人民政府2018年政府工作报告                                            ', '2018-03-30'), ('平昌县兰草镇人民政府', '兰草镇人民政府2017年政府工作报告                                            ', '2017-03-30'), ('平昌县兰草镇人民政府', '兰草镇人民政府2016年政府工作报告                                            ', '2016-10-31'), ('平昌县兰草镇人民政府', '兰草镇人民政府2015年政府工作报告                                            ', '2015-04-23'), ('平昌县兰草镇人民政府', '兰草镇人民政府2014年政府工作报告                                            ', '2014-04-01'), ('平昌县澌岸镇人民政府', '2022年工作总结和2023年工作计划                                            ', '2023-04-14'), ('平昌县澌岸镇人民政府', '澌岸镇2023年政府工作报告                                            ', '2023-02-08'), ('平昌县澌岸镇人民政府', '2022年政府工作报告                                            ', '2022-04-01'), ('平昌县澌岸镇人民政府', '澌岸镇2021年政府工作报告                                            ', '2021-10-23'), ('平昌县澌岸镇人民政府', '澌岸镇2020年政府工作报告                                            ', '2020-07-16'), ('平昌县澌岸镇人民政府', '澌岸镇人民政府2019年政府工作报告                                            ', '2019-03-21'), ('平昌县粉壁镇人民政府', '粉壁镇2023年政府工作报告                                            ', '2023-04-03'), ('平昌县粉壁镇人民政府', '粉壁镇2022年政府工作报告                                            ', '2022-04-07'), ('平昌县粉壁镇人民政府', '粉壁镇2021年政府工作报告                                            ', '2021-03-01'), ('平昌县粉壁镇人民政府', '粉壁镇人民政府2019年政府工作报告                                            ', '2019-03-28'), ('平昌县粉壁镇人民政府', '粉壁乡人民政府2018年政府工作报告                                            ', '2018-03-29'), ('平昌县粉壁镇人民政府', '粉壁乡人民政府2017年政府工作报告                                            ', '2017-03-29'), ('平昌县得胜镇人民政府', '得胜镇2023年政府工作报告                                            ', '2023-04-04'), ('平昌县得胜镇人民政府', '2021年政府工作报告                                            ', '2022-05-10'), ('平昌县得胜镇人民政府', '2022年政府工作报告                                            ', '2022-03-23'), ('平昌县得胜镇人民政府', '2021年政府工作报告                                            ', '2021-09-23'), ('平昌县得胜镇人民政府', '2020年得胜镇政府工作报告                                            ', '2020-09-23'), ('平昌县得胜镇人民政府', '得胜镇人民政府2019年政府工作报告                                            ', '2019-03-26'), ('平昌县得胜镇人民政府', '得胜镇人民政府2018年度政府工作报告                                            ', '2018-03-21'), ('平昌县得胜镇人民政府', '得胜镇人民政府2017年度政府工作报告                                            ', '2017-03-30'), ('平昌县元山镇人民政府', '元山镇人民政府2023年政府工作报告                                            ', '2023-03-31'), ('平昌县元山镇人民政府', '元山镇2022年政府工作报告                                            ', '2022-08-30'), ('平昌县元山镇人民政府', '元山镇2021年政府工作报告                                            ', '2021-12-13'), ('平昌县元山镇人民政府', '元山镇2020年政府工作报告                                            ', '2020-09-11'), ('平昌县元山镇人民政府', '元山镇2019年政府工作报告                                            ', '2020-09-11'), ('平昌县元山镇人民政府', '元山镇2018年政府工作报告                                            ', '2020-04-01'), ('平昌县元山镇人民政府', '元山镇2017年政府工作报告                                            ', '2020-04-01'), ('平昌县灵山镇人民政府', '灵山镇2023年政府工作报告                                            ', '2023-04-04'), ('平昌县灵山镇人民政府', '灵山镇2021年政府工作报告                                            ', '2022-08-17'), ('平昌县灵山镇人民政府', '灵山镇2022年政府工作报告                                            ', '2022-06-15'), ('平昌县灵山镇人民政府', '灵山镇2022年政府工作报告                                            ', '2022-06-13'), ('平昌县灵山镇人民政府', '灵山镇2020年政府工作报告                                            ', '2022-01-26'), ('平昌县灵山镇人民政府', '2019年政府工作报告                                            ', '2019-04-02'), ('平昌县灵山镇人民政府', '灵山镇2018年政府工作报告                                            ', '2018-03-28'), ('平昌县灵山镇人民政府', '灵山镇人民政府2017年政府工作报告                                            ', '2017-03-25'), ('平昌县土兴镇人民政府', '土兴镇人民政府2023年政府工作报告                                            ', '2023-04-19'), ('平昌县土兴镇人民政府', '2022年政府工作报告                                            ', '2022-06-16'), ('平昌县土兴镇人民政府', '土兴镇2021年政府工作报告                                            ', '2021-10-26'), ('平昌县土兴镇人民政府', '土兴镇2021年政府工作报告                                            ', '2021-02-26'), ('平昌县土兴镇人民政府', '土兴镇2020年政府工作报告                                            ', '2020-09-08'), ('平昌县土兴镇人民政府', '土兴镇2019年政府工作报告                                            ', '2019-04-19'), ('平昌县土兴镇人民政府', '土兴镇2018年政府工作报告                                            ', '2018-05-10'), ('平昌县土兴镇人民政府', '土兴镇人民政府2017年政府工作报告                                            ', '2017-04-06'), ('平昌县云台镇人民政府', '云台镇2023年政府工作报告                                            ', '2023-05-08'), ('平昌县云台镇人民政府', '云台镇2022年政府工作报告                                            ', '2022-06-22'), ('平昌县云台镇人民政府', '云台镇2021年政府工作报告                                            ', '2021-10-27'), ('平昌县云台镇人民政府', '云台镇2020年政府工作报告                                            ', '2020-09-15'), ('平昌县云台镇人民政府', '云台镇2019年政府工作报告                                            ', '2019-06-12'), ('平昌县云台镇人民政府', '云台镇2018年政府工作报告                                            ', '2018-03-28'), ('平昌县三十二梁镇人民政府', '三十二梁镇2023年政府工作报告                                            ', '2023-04-27'), ('平昌县三十二梁镇人民政府', '三十二梁镇2022年政府工作报告                                            ', '2022-07-20'), ('平昌县三十二梁镇人民政府', '三十二粱镇2021年政府工作报告                                            ', '2021-12-25'), ('平昌县三十二梁镇人民政府', '三十二梁镇2020年政府工作报告                                            ', '2020-10-06'), ('平昌县三十二梁镇人民政府', '石垭乡2019年政府工作报告                                            ', '2019-05-21'), ('平昌县板庙镇人民政府', '2023年政府工作报告                                            ', '2023-09-13'), ('平昌县板庙镇人民政府', '2022年政府工作报告                                            ', '2022-08-30'), ('平昌县板庙镇人民政府', '2021年政府工作报告                                            ', '2021-12-08'), ('平昌县板庙镇人民政府', '2020年政府工作安排报告                                            ', '2021-05-27'), ('平昌县板庙镇人民政府', '关于2019年政府工作报告的说明                                            ', '2020-04-24'), ('平昌县邱家镇人民政府', '邱家镇2023年政府工作报告                                            ', '2023-05-24'), ('平昌县邱家镇人民政府', '邱家镇人民政府2022年政府工作报告                                            ', '2022-06-16'), ('平昌县邱家镇人民政府', '邱家镇人民政府2021年政府工作报告                                            ', '2022-03-26'), ('平昌县邱家镇人民政府', '邱家镇人民政府2020年政府工作报告                                            ', '2021-08-06'), ('平昌县笔山镇人民政府', '笔山镇2022年法治政府建设报告                                            ', '2023-03-03'), ('平昌县笔山镇人民政府', '笔山镇2022年政府工作报告                                            ', '2022-07-05'), ('平昌县笔山镇人民政府', '2021年笔山镇人民政府工作报告                                            ', '2021-10-25'), ('平昌县笔山镇人民政府', '2020年平昌县笔山镇人民政府工作报告（草案）                                            ', '2020-09-14'), ('平昌县笔山镇人民政府', '2019年平昌县笔山镇人民政府工作报告（草案）                                            ', '2019-03-29'), ('平昌县笔山镇人民政府', '2018年平昌县笔山镇人民政府工作报告（草案）                                            ', '2018-04-02'), ('平昌县泥龙镇人民政府', '泥龙镇2023年政府工作报告                                            ', '2023-04-18'), ('平昌县泥龙镇人民政府', '泥龙镇2022年政府工作报告                                            ', '2022-04-12'), ('平昌县泥龙镇人民政府', '泥龙镇2021年政府工作报告                                            ', '2021-08-16'), ('平昌县泥龙镇人民政府', '泥龙镇2020年度政府工作报告                                            ', '2020-10-13'), ('平昌县泥龙镇人民政府', '2017年政府工作报告                                            ', '2018-06-19'), ('平昌县泥龙镇人民政府', '泥龙镇2018年政府工作报告                                            ', '2018-04-17'), ('平昌县岩口镇人民政府', '岩口镇2023年政府工作报告                                            ', '2023-04-18'), ('平昌县岩口镇人民政府', '岩口镇2022年政府工作报告                                            ', '2022-04-12'), ('平昌县岩口镇人民政府', '岩口镇2021年政府工作报告                                            ', '2021-10-26'), ('平昌县岩口镇人民政府', '岩口乡2019年政府工作报告                                            ', '2020-04-08'), ('平昌县岩口镇人民政府', '岩口乡2017年政府工作报告                                            ', '2020-04-08'), ('平昌县岩口镇人民政府', '岩口镇人民政府2018年政府工作报告                                            ', '2018-03-30'), ('平昌县镇龙镇人民政府', '平昌县镇龙镇2023年政府工作报告                                            ', '2023-03-30'), ('平昌县镇龙镇人民政府', '镇龙镇2022年政府工作报告                                            ', '2022-08-08'), ('平昌县镇龙镇人民政府', '镇龙镇2021年政府工作报告                                            ', '2021-08-17'), ('平昌县镇龙镇人民政府', '镇龙镇2020年政府工作报告                                            ', '2020-08-25'), ('平昌县镇龙镇人民政府', '镇龙镇2018年政府工作报告                                            ', '2018-11-02'), ('平昌县镇龙镇人民政府', '镇龙镇2017年政府工作报告                                            ', '2017-05-26'), ('平昌县镇龙镇人民政府', '镇龙镇2015年政府工作报告                                            ', '2015-05-11'), ('平昌县镇龙镇人民政府', '镇龙镇2013年政府工作报告                                            ', '2013-08-06'), ('平昌县望京镇人民政府', '望京镇人民政府2023年政府工作报告                                            ', '2023-04-20'), ('平昌县望京镇人民政府', '望京镇人民政府2022年政府工作报告                                            ', '2022-04-06'), ('平昌县望京镇人民政府', '望京镇人民政府2021年政府工作报告                                            ', '2021-03-03'), ('平昌县望京镇人民政府', '望京镇人民政府2020年政府工作报告                                            ', '2020-10-14'), ('平昌县望京镇人民政府', '望京镇人民政府2019年政府工作报告                                            ', '2019-02-05'), ('平昌县望京镇人民政府', '望京镇人民政府2018年政府工作报告                                            ', '2018-03-06'), ('平昌县望京镇人民政府', '望京镇人民政府2017年政府工作报告                                            ', '2017-03-02'), ('平昌县望京镇人民政府', '望京镇人民政府2016年政府工作报告                                            ', '2016-03-31'), ('平昌县望京镇人民政府', '望京镇人民政府2015年政府工作报告                                            ', '2015-04-01'), ('平昌县江家口镇人民政府', '江家口镇人民政府2023年政府工作报告                                            ', '2023-08-17'), ('平昌县江家口镇人民政府', '江家口镇2022年政府工作报告                                            ', '2022-08-09'), ('平昌县江家口镇人民政府', '江家口镇2021年政府工作报告                                            ', '2021-10-28'), ('平昌县江家口镇人民政府', '2020年政府工作建议                                            ', '2020-09-15')]
    makeExcel(zf_work_list)
if __name__ == '__main__':

    startMain()




