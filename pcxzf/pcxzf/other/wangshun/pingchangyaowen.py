import requests
import conf
import datetime
import openpyxl
import collections
import os
import sys
import tempfile
import platform


def get_pcyw():
    page_index = 1
    pcyw_news_list = []
    while True:
        # 构造请求的URL
        url = f'http://www.scpc.gov.cn/content/column/6789791?pageIndex={page_index}'
        # 尝试发送请求并获取响应
        try:
            response = conf.make_request_get(url)
        # 如果发生错误，打印错误信息并退出循环
        except Exception as e:
            print(f'Error: {e}')
            break

        if not deal_res(response, pcyw_news_list,None):
            break

        page_index += 1
    print(pcyw_news_list)
    return pcyw_news_list


def get_bmgz(pcyw_list):
    page_index = 1
    bmgz_news_list = []
    while True:
        # 构造请求的URL
        url = f'http://www.scpc.gov.cn/content/column/6789801?pageIndex={page_index}'
        # 尝试发送请求并获取响应
        try:
            response = conf.make_request_get(url)
        # 如果发生错误，打印错误信息并退出循环
        except Exception as e:
            print(f'Error: {e}')
            break

        if not deal_res(response, bmgz_news_list,pcyw_list):
            break

        page_index += 1
    print(bmgz_news_list)
    return bmgz_news_list
def get_jcdt(pcyw_list):
    page_index = 1
    jcdt_news_list = []
    while True:
        # 构造请求的URL
        url = f'http://www.scpc.gov.cn/content/column/6789811?pageIndex={page_index}'
        # 尝试发送请求并获取响应
        try:
            response = conf.make_request_get(url)
        # 如果发生错误，打印错误信息并退出循环
        except Exception as e:
            print(f'Error: {e}')
            break

        if not deal_res(response, jcdt_news_list, pcyw_list):
            break

        page_index += 1
    print(jcdt_news_list)
    return jcdt_news_list


def startMain():

    pingchangyaowen_list = get_pcyw()

    bumengongzuo_list = get_bmgz(pingchangyaowen_list)

    jcdt_list = get_jcdt(pingchangyaowen_list)

    make_excel(pingchangyaowen_list, conf.pcyw_excel_name)
    make_excel(bumengongzuo_list, conf.bmgz_excel_name)
    make_excel(jcdt_list, conf.jcdt_excel_name)


def make_excel(news_list,excelname):
    # 创建一个Workbook对象
    wb = openpyxl.Workbook()

    # 获取当前的Worksheet对象
    ws = wb.active

    # 遍历列表中的每一个元组
    for tup in news_list:
        # 将元组添加到表格的一行中
        ws.append(tup)
        print(tup)
    counter = collections.Counter(tup[2] for tup in news_list)

    # 打印计数结果
    print(counter)
    ws.append(['单位', '数量'])

    # 遍历字典中的每一个键值对
    for key, value in counter.items():
        # 将键值对添加到表格的一行中
        ws.append([key, value])

    # 保存表格到文件中
    wb.save(excelname)
def jugedate(date):
    date_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
    date_target = datetime.datetime.strptime(conf.target_date, "%Y-%m-%d")
    return date_obj < date_target
def deal_news_href(href):
    if not href.startswith('http://www.scpc.gov.cn'):
        return 1,""
    res_dup = conf.make_request_get(href)

    div = res_dup[1].find('div', class_='newsinfo')
    if div is None:
        return None
    info = {span.text.split('：')[0]: span.text.split('：')[1] for span in div.find_all('span')[:3] if '：' in span.text}

    if '来源' in info:
        # 判断info['来源']是否包含out_comp列表中的任意一个字段
        if not any(field in info['来源'] for field in conf.out_comp):
            # 如果不包含，判定为JIA
            return 0,info['来源'],info['作者']  if '作者' in info else ""
    return 1,""

def deal_tittle(str,pcyw_list):
    tittle = str.split('【')[-1].split('】')[-1]
    if pcyw_list is not None:
        for tup in pcyw_list:
            # 判断字符串是否在元组的第一个元素中
            if tittle in tup[0]:
                return 1,''
    return  0,tittle
def jugedate_after(date):
    date_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
    after_date = datetime.datetime.strptime(conf.after_date, "%Y-%m-%d")
    return date_obj >= after_date
def deal_res(response, news_list,pcyw_list):
# 这里返回假则不再翻页
    div_elment = response[1].find('div', {'class': 'listnews'})
    if div_elment is not None:
        li_elment = div_elment.find_all('li')
        for li in li_elment:
            if li.find('span', class_='right date') is not None:
                date = li.find('span', class_='right date').text
                if  jugedate(date):
                    return False
                if jugedate_after(date):
                    continue
                title_deal = deal_tittle(li.find('a', class_='left')['title'],pcyw_list)
                if title_deal[0]:
                    continue
                href = li.find('a', class_='left')['href']
                newsinfo = deal_news_href(href)
                if newsinfo[0]:
                    continue
                news_list.append((title_deal[1],date,newsinfo[1],newsinfo[2]))

    # 第一页测试结束这里改为真

    return True

def test_excel_writing():
    """
    测试Excel文件写入功能，检查各种可能的问题
    """
    print("\n===== Excel写入测试开始 =====")
    print(f"当前时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Python版本: {sys.version}")
    print(f"操作系统: {platform.platform()}")
    print(f"当前工作目录: {os.getcwd()}")
    try:
        print(f"当前用户: {os.getlogin()}")
    except Exception as e:
        print(f"获取用户名失败: {e}")
    
    # 测试1: 在当前目录写入
    print("\n测试1: 在当前目录写入")
    current_dir_file = "test_current_dir.xlsx"
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["测试", "数据"])
        ws.append(["当前目录", "写入测试"])
        print(f"尝试保存到当前目录: {os.path.abspath(current_dir_file)}")
        wb.save(current_dir_file)
        if os.path.exists(current_dir_file):
            print(f"✅ 成功: 文件已保存到当前目录，大小: {os.path.getsize(current_dir_file)} 字节")
        else:
            print("❌ 失败: 文件未找到，但无异常抛出")
    except Exception as e:
        print(f"❌ 异常: {e}")
    
    # 测试2: 在临时目录写入
    print("\n测试2: 在临时目录写入")
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, "test_temp_dir.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["测试", "数据"])
        ws.append(["临时目录", "写入测试"])
        print(f"尝试保存到临时目录: {temp_file}")
        wb.save(temp_file)
        if os.path.exists(temp_file):
            print(f"✅ 成功: 文件已保存到临时目录，大小: {os.path.getsize(temp_file)} 字节")
        else:
            print("❌ 失败: 文件未找到，但无异常抛出")
    except Exception as e:
        print(f"❌ 异常: {e}")
    
    # 测试3: 在用户目录写入
    print("\n测试3: 在用户目录写入")
    user_dir = os.path.expanduser("~")
    user_file = os.path.join(user_dir, "test_user_dir.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["测试", "数据"])
        ws.append(["用户目录", "写入测试"])
        print(f"尝试保存到用户目录: {user_file}")
        wb.save(user_file)
        if os.path.exists(user_file):
            print(f"✅ 成功: 文件已保存到用户目录，大小: {os.path.getsize(user_file)} 字节")
        else:
            print("❌ 失败: 文件未找到，但无异常抛出")
    except Exception as e:
        print(f"❌ 异常: {e}")
    
    # 测试4: 检查文件是否被占用
    print("\n测试4: 检查文件占用情况")
    occupied_file = "test_occupied.xlsx"
    try:
        # 创建文件
        wb1 = openpyxl.Workbook()
        wb1.save(occupied_file)
        print(f"创建文件: {occupied_file}")
        
        # 尝试打开并保持文件句柄
        with open(occupied_file, "rb") as f:
            print("文件已打开并保持句柄")
            
            # 尝试在文件被打开的情况下重新写入
            try:
                wb2 = openpyxl.Workbook()
                print("尝试写入被占用的文件...")
                wb2.save(occupied_file)
                print("✅ 成功: 即使文件被打开，依然可以写入")
            except Exception as e:
                print(f"❌ 预期的异常: {e}")
        
        # 清理
        if os.path.exists(occupied_file):
            os.remove(occupied_file)
            print(f"测试文件已清理: {occupied_file}")
    except Exception as e:
        print(f"测试过程中出现异常: {e}")
    
    # 测试5: 总结你的代码中的情况
    print("\n测试5: 模拟你的make_excel方法")
    sample_data = [
        ("标题1", "内容1", "来源1"),
        ("标题2", "内容2", "来源2"),
        ("标题3", "内容3", "来源1")
    ]
    test_file = "test_make_excel.xlsx"
    try:
        print(f"调用make_excel函数，保存到: {os.path.abspath(test_file)}")
        make_excel(sample_data, test_file)
        if os.path.exists(test_file):
            print(f"✅ 成功: make_excel函数正常工作，文件大小: {os.path.getsize(test_file)} 字节")
        else:
            print("❌ 失败: make_excel未能创建文件")
    except Exception as e:
        print(f"❌ make_excel函数异常: {e}")
    
    print("\n===== Excel写入测试结束 =====")
    print("如果看到多个'成功'消息，但你的实际代码仍然不工作，可能是:")
    print("1. 文件路径问题: 你的代码可能保存到了其他位置")
    print("2. 权限问题: 不同运行环境的权限不同")
    print("3. 文件占用: 目标文件可能被其他程序占用")
    print(f"测试文件保存在以下位置:")
    print(f"- 当前目录: {os.path.abspath(current_dir_file)}")
    print(f"- 临时目录: {temp_file}")
    print(f"- 用户目录: {user_file}")
    print(f"- 模拟测试: {os.path.abspath(test_file)}")
    

if __name__ == '__main__':
    test_excel_writing()
    # 其他代码...


