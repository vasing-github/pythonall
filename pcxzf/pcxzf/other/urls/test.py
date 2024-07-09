import openpyxl

# 创建一个新的工作簿
import requests

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "URLs"

# 定义数据
data = {
    '公示公告': 'http://www.scpc.gov.cn/ywdt/gsgg/index.html',
    '基层动态': 'http://www.scpc.gov.cn/ywdt/xzdt/index.html',
    '部门工作': 'http://www.scpc.gov.cn/ywdt/bmdt/index.html',
    '热点关注': 'http://www.scpc.gov.cn/ywdt/rdgz/index.html',
    '便民信息': 'http://www.scpc.gov.cn/ywdt/bmxx/index.html',
    '平昌要闻': 'http://www.scpc.gov.cn/ywdt/pcyw/index.html',
    '政策解读': 'http://www.scpc.gov.cn/zwgk/zcjd/index.html',
    '处罚强制': 'http://www.scpc.gov.cn/zwgk/jbxxgk/cfqz/index.html',
    '财政预决算': 'http://www.scpc.gov.cn/ztzl/pcxczyjsgk/czyjs/index.html',
    '部门预算': 'http://www.scpc.gov.cn/ztzl/pcxczyjsgk/bmys/index.html',
    '部门决算': 'http://www.scpc.gov.cn/ztzl/pcxczyjsgk/bmjs/index.html',
    '单位预算': 'http://www.scpc.gov.cn/ztzl/pcxczyjsgk/dwys/index.html',
    '单位决算': 'http://www.scpc.gov.cn/ztzl/pcxczyjsgk/dwjs/index.html',
    '行政事业性收费': 'http://www.scpc.gov.cn/zwgk/jbxxgk/xzsyxsf/index.html',
    '县委常委会': 'http://www.scpc.gov.cn/zwgk/xwcwh/index.html',
    '县政府常务会': 'http://www.scpc.gov.cn/zwgk/xzfcwh/index.html',
    '全体会议': 'http://www.scpc.gov.cn/zwgk/jbxxgk/qthy/index.html',
    '人事信息': 'http://www.scpc.gov.cn/zwgk/rsxx/index.html',
    '统计信息': 'http://www.scpc.gov.cn/zwgk/jbxxgk/tjxx/index.html',
    '意见征集': 'http://www.scpc.gov.cn/zwgk/jbxxgk/jcygk/yjzj/index.html',
    '结果反馈': 'http://www.scpc.gov.cn/zwgk/jbxxgk/jcygk/jgfk/index.html',
    '决策结果': 'http://www.scpc.gov.cn/zwgk/jbxxgk/jcygk/jcjg/index.html',
    '助企纾困': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/zqsk/index.html',
    '减税降费': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/jsjf/index.html',
    '公共资源交易': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/zyjy/index.html',
    '水质监测': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/hjbh/index.html',
    '征地信息': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/zdxx/index.html',
    '价格信息': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/jghsf/index.html',
    '不动产登记': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/bdcdj/index.html',
    '财政资金直达基层': 'http://www.scpc.gov.cn/zwgk/zdlyxxgk/czzjzdjc/index.html',
    '卫生健康': 'http://www.scpc.gov.cn/public/column/6602161?type=4&action=list&nav=2&sub=6&catId=6715931',
    '食品药品监管': 'http://www.scpc.gov.cn/public/column/6602241?type=4&action=list&nav=2&sub=6&catId=6715931',
    '就业创业': 'http://www.scpc.gov.cn/public/column/6602001?type=4&action=list&nav=2&sub=5&catId=6715921',
    '养老服务': 'http://www.scpc.gov.cn/public/column/6601941?type=4&catId=6715921&action=list',
    '教育科技': 'http://www.scpc.gov.cn/public/column/6601901?type=4&action=list&nav=2&sub=5&catId=6715921',
    '涉农补贴': 'http://www.scpc.gov.cn/public/column/6602101?type=4&action=list&nav=2&sub=4&catId=6715911',
    '公共文化服务': 'http://www.scpc.gov.cn/public/column/6602141?type=4&action=list&nav=2&sub=4&catId=6715921',
    '社会救助': 'http://www.scpc.gov.cn/public/column/6601941?type=4&action=list&nav=2&sub=6&catId=6715931',
    '住房保障': 'http://www.scpc.gov.cn/public/column/6602041?type=4&action=list&nav=2&sub=6&catId=6715931',
    '放管服改革': 'http://www.scpc.gov.cn/public/column/6602381?type=4&action=list&nav=2&sub=5&catId=6715931',
    '应急管理': 'http://www.scpc.gov.cn/public/column/6602201?type=4&action=list&nav=2&sub=5&catId=6715921',
    '环评公示': 'http://www.scpc.gov.cn/public/column/6602261?type=4&action=list&nav=2&sub=4&catId=6715911',
    '监测预警': 'http://www.scpc.gov.cn/public/column/6602201?type=4&action=list&nav=2&sub=5&catId=6715931',
    '医疗保障': 'http://www.scpc.gov.cn/public/column/6602361?type=4&action=list&nav=2&sub=5&catId=6715931',
    '环境质量公报': 'http://www.scpc.gov.cn/public/column/6602261?type=4&action=list&nav=2&sub=6&catId=6715931',
    '两会专题': 'http://www.scpc.gov.cn/ztzl/lhzt/index.html',
    '乡村振兴': 'http://www.scpc.gov.cn/ztzl/tpbk/index.html',
    '依法治县': 'http://www.scpc.gov.cn/ztzl/yfzx/index.html',
    '镇街政府信息主动公开目录': 'http://www.scpc.gov.cn/ztzl/zfxxzdgkjbml/xzbsc/index.html',
    '县级部门政府信息主动公开目录': 'http://www.scpc.gov.cn/ztzl/zfxxzdgkjbml/xjbm/index.html',
    '行政许可': 'http://www.scpc.gov.cn/ztzl/xyxxsgs/xzxk/index.html',
    '行政处罚': 'http://www.scpc.gov.cn/ztzl/xyxxsgs/xzcf/index.html',
    '双公示目录': 'http://www.scpc.gov.cn/ztzl/xyxxsgs/xkml/index.html',
    '拖底帮扶': 'http://www.scpc.gov.cn/ztzl/wqqxkpctdxbfjxs/index.html',
    '部门试点目录': 'http://www.scpc.gov.cn/ztzl/pcxsdlyjczwgkbzml/pcxbmsdlyjczwgkbzml/index.html',
    '乡镇试点目录': 'http://www.scpc.gov.cn/ztzl/pcxsdlyjczwgkbzml/pcxgzjdbscsdlyjczwgkbzml/index.html',
}

# # 写入数据到工作表
# for row, (name, url) in enumerate(data.items(), start=1):
#     ws.cell(row=row, column=1, value=name)
#     ws.cell(row=row, column=2, value=url)
#
# # 保存工作簿
# wb.save("urls.xlsx")

# 定义请求头
headers = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'Authorization': 'Bearer 4a988595-036f-4f74-982b-e446a9064269',
    'Connection': 'keep-alive',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36 Edg/128.0.0.0',
    'sec-ch-ua': '"Chromium";v="128", "Not;A=Brand";v="24", "Microsoft Edge";v="128"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"'
}

# 定义检查URL是否存在的函数
def check_url_exists(url):
    check_url = f'https://datais.ucap.com.cn/cloud-website-web/websiteChannelUpdate/checkChannelAddress?url={url}'
    response = requests.get(check_url, headers=headers)
    return response.json().get('data', False)

# 定义添加URL的函数
def add_url(name, url):
    add_url = 'https://datais.ucap.com.cn/cloud-website-web/websiteChannelUpdate/insertSaveDto'
    payload = {
        "channelName": name,
        "channelAddress": url,
        "firstClassificationId": 1,
        "secondaryClassificationId": 11,
        "isUpdate": 1,
        "protectedLevel": 0,
        "siteCode": "5119230005"
    }
    response = requests.post(add_url, headers=headers, json=payload)
    return response.json()

# 遍历数据字典并执行操作
for name, url in data.items():
    if not check_url_exists(url):
        result = add_url(name, url)
        print(f"Added {name}: {result}")
    else:
        print(f"{name} already exists, skipping.")

# # 写入数据到工作表
# for row, (name, url) in enumerate(data.items(), start=1):
#     ws.cell(row=row, column=1, value=name)
#     ws.cell(row=row, column=2, value=url)
#
# # 保存工作簿
# wb.save("urls.xlsx")
