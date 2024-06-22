import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook

def putin():
    # 数据库连接配置
    db_config = {
           "host": "10.167.39.125",
    "user": "kfvasing",
    "password": "vas9624..",
    "database": "smartpingchangdb",
        'charset': 'utf8mb4'
    }

    # 连接到数据库
    connection = pymysql.connect(**db_config)


    # 打开Excel文件
    workbook = load_workbook('project.xlsx')
    sheet = workbook.active

    # 准备SQL插入语句
    # 准备新的SQL插入语句
    insert_stmt = (
        "INSERT INTO pc_investment (project_name, project_status, project_stage, "
        "project_urgent, status, contact_name, contact_phone, contact_position, "
        "company_name, company_introduce, company_adress, company_regist, "
        "company_nature, company_tag, company_main, company_industry, company_area, "
        "staff_communicate, input_creat_time, inveestment_text, inveestment_money, "
        "project_contact, company_need, question, suguesstion, main_contact_department, "
        "coordinate_department1, coordinate_department2, coordinate_department3, "
        "staff_creat_time, contact_leader, import_creat_time, sign_build_text, "
        "sign_local_name, in_overall_money, sign_time, start_build_time, "
        "in_overall_time, sign_push, sign_creat_time, other, other1, other2, "
        "other3, extra1, extra2, extra3, extra4, extra5, extra6, extra7) "
        "VALUES (%s, %s,%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    )


    # 从第二行开始遍历（假设第一行是列头）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 将Excel中的空单元格转换为None
        row = [str(cell) if isinstance(cell, (int, float)) else cell for cell in row]
        row = [None if cell is None or str(cell).strip() == '' else cell for cell in row]
        # 执行SQL语句
        print(insert_stmt)
        print(row)
        try:
            with connection.cursor() as cursor:
                cursor.execute(insert_stmt, row)
            connection.commit()
        except Exception as e:
            print("发生错误：", e)
            print("SQL语句：", insert_stmt)
            print("当前行数据：", row)
            # 如果发生错误，回滚事务
            connection.rollback()



    # 关闭数据库连接
    connection.close()


def createxcel():


    # 创建一个新的工作簿
    workbook = Workbook()
    sheet = workbook.active

    # 定义第一行的字段
    headers = [
        "project_name", "project_status", "project_stage",
        "project_urgent", "status", "contact_name", "contact_phone", "contact_position",
        "company_name", "company_introduce", "company_adress", "company_regist",
        "company_nature", "company_tag", "company_main", "company_industry", "company_area",
        "staff_communicate", "input_creat_time", "inveestment_text", "inveestment_money",
        "project_contact", "company_need", "question", "suguesstion", "main_contact_department",
        "coordinate_department1", "coordinate_department2", "coordinate_department3",
        "staff_creat_time", "contact_leader", "import_creat_time", "sign_build_text",
        "sign_local_name", "in_overall_money", "sign_time", "start_build_time",
        "in_overall_time", "sign_push", "sign_creat_time", "other", "other1", "other2",
        "other3", "extra1", "extra2", "extra3", "extra4", "extra5", "extra6", "extra7"
    ]

    # 将字段填入第一行
    sheet.append(headers)

    # 保存工作簿
    workbook.save("project.xlsx")


if __name__ == '__main__':
    # createxcel()
    putin()