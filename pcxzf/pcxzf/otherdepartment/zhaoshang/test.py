import pymysql
from openpyxl import load_workbook

# 连接到数据库
connection = pymysql.connect(host='yuxiaohaishidalao.shop', user='root', password='vas9624..', db='smartpc', charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)

try:
    with connection.cursor() as cursor:
        # 打开 Excel 文件
        workbook = load_workbook(filename='chubei.xlsx')
        sheet = workbook.active

        # 遍历表格中的每一行
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是列名
            # 检查联系人信息是否为空
            contact_info = row[3] if row[3] else ''
            if contact_info:
                parts = contact_info.split(',')  # 假设姓名和电话号码用逗号分隔
                contact_name = parts[0].strip()
                contact_phone = parts[1].strip() if len(parts) > 1 else ''

                # 插入数据到数据库
                sql = """
                       INSERT INTO `pc_investment` (`project_name`, `company_name`, `contact_name`, `contact_phone`, `company_introduce`, `company_need`, `inveestment_text`, `inveestment_money`, `project_contact`, `question`, `suguesstion`, `contact_department`, `contact_leader`, `project_status`, `other`)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, '0', %s)
                       """
                print(sql)
                cursor.execute(sql, (
                row[1], row[2], contact_name, contact_phone, row[4], row[8], row[5], row[6], row[7], row[9], row[10],
                row[12], row[11], row[14]))

    # 提交事务
    connection.commit()

finally:
    connection.close()
