import pymysql
from openpyxl import load_workbook

# 数据库连接配置
db_config = {
    'host': 'yuxiaohaishidalao.shop',
    'user': 'root',
    'password': 'vas9624..',
    'database': 'smartpc',
    'charset': 'utf8mb4'
}

# 连接到数据库
connection = pymysql.connect(**db_config)


# 打开Excel文件
workbook = load_workbook('product.xlsx')
sheet = workbook.active

# 准备SQL插入语句
insert_stmt = (
    "INSERT INTO pc_jk_product_info (company_name, product_name, product_standard, "
    "product_price, product_img, product_output_overal, product_output_max_month, "
    "contact_method, product_zizhi_img, product_honor_img, other, extra1, extra2, "
    "extra3, extra4, extra5, extra6, extra7) "
    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
)


# 从第二行开始遍历（假设第一行是列头）
for row in sheet.iter_rows(min_row=2, values_only=True):
    # 将Excel中的空单元格转换为None
    row = [None if cell is None or str(cell).strip() == '' else cell for cell in row]
    # 执行SQL语句
    print(insert_stmt)
    print(row)
    with connection.cursor() as cursor:
        cursor.execute(insert_stmt, row)
    # 提交事务
    connection.commit()

# 关闭数据库连接
connection.close()