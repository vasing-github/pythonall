# -*- coding: utf-8 -*-
import json
import os
import re
import sys
from datetime import datetime, timedelta
from io import BytesIO
import hudongcontent
import requests
current_dir = os.path.dirname(os.path.abspath(__file__))
# 获取项目的根目录
project_root = os.path.abspath(os.path.join(current_dir, os.pardir, os.pardir))
# 将项目根目录添加到 sys.path
sys.path.append(project_root)
from kaipumodify.cfg import dealtext
from kaipumodify.cfg import conf
import kaipumodify.cfg.text as text
from openpyxl import Workbook
from openpyxl import load_workbook
import getcontent2, getcontent

token = text.token

bz_gov_id = text.bz_gov_id
jid = text.jid

endDate = datetime.now()
startDate = endDate - timedelta(days=30)
endDate_str = endDate.strftime('%Y-%m-%d')
startDate_str = startDate.strftime('%Y-%m-%d')


def get_cuomin_list():
    cookies = {
        'HWWAFSESID': '49d6373a41e52c2bd8',
        'HWWAFSESTIME': '1716168884097',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=49d6373a41e52c2bd8; HWWAFSESTIME=1716168884097',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0',
        'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Microsoft Edge";v="126"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 100,
            'current': 1,
        },
        'check': 1,
        'dataState': True,
        'dateType': 1,
        'codeType': 1,
        'custCode': '5119230005',
        'startDate': startDate_str,
        'endDate': endDate_str,
        'rectifyStatus': 0,
        'searchBox': '',
        'questionLevel': '',
        'labelIds': [],
        'resultType': '',
        'pageType': '',
        'recommendUpdateList': [],
        'protectCode': '5119230005',
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/listPageByDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print(response.text)
    return response.json()['data']['records']


def get_link_use_num(token):
    cookies = {
        'HWWAFSESID': 'b092d732d20af40cfc',
        'HWWAFSESTIME': '1709105903507',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=b092d732d20af40cfc; HWWAFSESTIME=1709105903507',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
        'sec-ch-ua': '"Microsoft Edge";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'startDate': startDate_str,
        'endDate': endDate_str,
        'status': 1,
        'dateType': 0,
        'isSure': 0,
        'urlType': '',
        'checkType': '',
        'siteInfo': '',
        'siteCustCode': '5119230005',
        'protectCode': '5119230005',
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteDeadLink/getStatisticsByQueryDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )

    a = response.json()['data']['channelNotRectified']
    b = response.json()['data']['homeNotRectified']
    print("link_use_num:", a + b)
    return a + b


def get_out_link(token):
    cookies = {
        'HWWAFSESID': 'b092d732d20af40cfc',
        'HWWAFSESTIME': '1709105903507',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=b092d732d20af40cfc; HWWAFSESTIME=1709105903507',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
        'sec-ch-ua': '"Microsoft Edge";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 100,
            'current': 1,
        },
        'check': 1,
        'dataState': True,
        'dateType': 4,
        'inspectionType': 1,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'codeType': 1,
        'codeTypeVal': [
            '5119230005',
        ],
        'pageType': '',
        'reviewType': '',
        'protectCode': '5119230005',
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteOutLink/reviewDetailStatistics',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print("out_link_num:", response.json()['data']['detail']['0'])
    return response.json()['data']['detail']['0']


def get_yinsi_num(token):
    cookies = {
        'HWWAFSESID': 'b092d732d20af40cfc',
        'HWWAFSESTIME': '1709105903507',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=b092d732d20af40cfc; HWWAFSESTIME=1709105903507',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
        'sec-ch-ua': '"Microsoft Edge";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'check': 1,
        'codeType': 1,
        'codeTypeVal': [
            '5119230005',
        ],
        'resultTypeList': [
            1,
            2,
        ],
        'dateType': 1,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'page': {
            'size': 100,
            'current': 1,
        },
        'pageType': '',
        'reviewType': 0,
        'protectCode': '5119230005',
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteInfoLeakageMaster/reviewStatistics',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print("yinsi_num:", response.json()['data']['allCount'])
    return response.json()['data']['allCount']


def start_search():
    print(endDate_str)
    cookies = {
        'HWWAFSESID': '40854762319bc867f7',
        'HWWAFSESTIME': '1703639525975',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=40854762319bc867f7; HWWAFSESTIME=1703639525975',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 10,
            'current': 1,
        },
        'check': 1,
        'codeType': 1,
        'dataState': True,
        'custCode': '5119230005',
        'dateType': 1,
        'startDate': startDate_str,
        'endDate': endDate_str,
        'rectifyStatus': '',
        'searchBox': '',
        'protectCode': '5119230005',
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/getRectifiedNum',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print(response.json())
    if response.status_code != 200:
        dealtoken()
        return start_search()
    waitrefix = response.json()['data']['notRectifiedNum']
    print('waitrefix:', waitrefix)

    return waitrefix


def dealtoken():
    global token
    token = dealtext.deal_kaipu_token()


def send_msg(cuomin, yinsi_num, outnum, kai_link_use_num, uptime=None):
    # 定义要发送的消息内容
    # 获取当前时间，精确到分钟
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" 检测到开普云有新错敏<font color=\"warning\">{cuomin}</font>条\n\n隐私泄露<font color=\"warning\">{yinsi_num}</font>条\n\n外链暗链<font color=\"warning\">{outnum}</font>条\n\n链接不可用<font color=\"warning\">{kai_link_use_num}</font>条\n\n\n开普检测时间：{uptime}\n\n机器人检测时间：{current_time}\n<@WuXiaoLong>\n"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def get_kaipu_uptime(token):
    cookies = {
        'HWWAFSESID': '40854762319bc867f7',
        'HWWAFSESTIME': '1703639525975',
    }
    endDate = datetime.now()
    startDate = endDate - timedelta(days=30)

    # 格式化日期
    endDate_str = endDate.strftime('%Y-%m-%d')
    startDate_str = startDate.strftime('%Y-%m-%d')
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=40854762319bc867f7; HWWAFSESTIME=1703639525975',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'page': {
            'size': 10,
            'current': 1,
        },
        'check': 1,
        'dataState': True,
        'dateType': 1,
        'codeType': 1,
        'custCode': '5119230005',
        'startDate': startDate_str,
        'endDate': endDate_str,
        'rectifyStatus': 0,
        'searchBox': '',
        'questionLevel': '',
        'labelIds': [],
        'resultType': '',
        'pageType': '',
        'recommendUpdateList': [],
        'protectCode': '5119230005',
        'custLevel': '1',
        'unitLevel': 3,
        'isHandoff': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/listPageByDto',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print(token)
    print(response.json()['data']['records'][0]['updateTime'])
    return response.json()['data']['records'][0]['updateTime']


def kaipucorrect(ids):
    cookies = {
        'HWWAFSESID': '49d6373a41e52c2bd8',
        'HWWAFSESTIME': '1716168884097',
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Authorization': 'Bearer ' + token,
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        # 'Cookie': 'HWWAFSESID=49d6373a41e52c2bd8; HWWAFSESTIME=1716168884097',
        'Origin': 'https://datais.ucap.com.cn',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0',
        'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Microsoft Edge";v="126"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'ids': ids,
        'rectifyStatus': 1,
    }

    response = requests.post(
        'https://datais.ucap.com.cn/cloud-website-web/websiteSensitiveDetail/updateCheckTypeByIds',
        cookies=cookies,
        headers=headers,
        json=json_data,
    )
    print(response.text)


def get_new_bzid_jid():
    global bz_gov_id, jid
    # print("bz_gov_id token press...........")
    # new_bz_gov_id, new_jid = getjsseion.get_new_cookie()
    # modify_cookie(new_bz_gov_id, new_jid)
    # save_data()
    bz_gov_id, jid = dealtext.jiyue_token()


def add_to_correct(correctlist, correctids, cuomin):
    correctids.append(cuomin['id'])
    correctlist.append((cuomin['sensitiveWords'], cuomin['recommendUpdate'], cuomin['snapshotNew'], cuomin['url'],
                        cuomin['parentUrl'], cuomin['pageTypeMeaning'], cuomin['parentTitle']))


def cuo_1_argument(numbers, cuomin, correctlist, correctids):
    res = getcontent2.getcontent(numbers[0], bz_gov_id, jid)
    if res['status'] == -9:
        get_new_bzid_jid()
        res = getcontent2.getcontent(numbers[0], bz_gov_id, jid)
    getcontent2.savearticnews(res, cuomin['sensitiveWords'], cuomin['recommendUpdate'], bz_gov_id,
                              jid)
    add_to_correct(correctlist, correctids, cuomin)


def cuo_2_aruments(numbers, cuomin, correctlist, correctids):
    res = getcontent.getcontent(numbers[0], numbers[1], bz_gov_id, jid)
    if res['status'] == -9:
        get_new_bzid_jid()
        res = getcontent.getcontent(numbers[0], numbers[1], bz_gov_id, jid)
    getcontent.saveorupdate(res, cuomin['sensitiveWords'], cuomin['recommendUpdate'], bz_gov_id,
                            jid)
    add_to_correct(correctlist, correctids, cuomin)


def cuo_hudong(cuomin, correctlist, correctids):
    is_need_up_toke = hudongcontent.start_kaipu(cuomin, bz_gov_id, jid)
    if is_need_up_toke:
        get_new_bzid_jid()
        hudongcontent.start_kaipu(cuomin, bz_gov_id, jid)
    add_to_correct(correctlist, correctids, cuomin)


def dealcuo():
    kaipu_waitrefix = start_search()
    if kaipu_waitrefix != 0:

        list_cuomin = get_cuomin_list()
        correctlist = []
        correctids = []
        for cuomin in list_cuomin:
            if cuomin['pageType'] == "3":  # 表示是文章类型
                if cuomin['column'] != '互动交流':
                    numbers = re.findall(r'\d+', cuomin['url'])
                    # 将提取出的数字转换为整数
                    numbers = [int(num) for num in numbers]
                    # 判断 URL 类型并返回结果
                    if len(numbers) == 1:
                        cuo_1_argument(numbers, cuomin, correctlist, correctids)
                    elif len(numbers) == 2:
                        cuo_2_aruments(numbers, cuomin, correctlist, correctids)
                    else:
                        print("未知情况")
                        send_nosee()


                elif cuomin['column'] == '互动交流':
                    cuo_hudong(cuomin, correctlist, correctids)

        send_correct_msg(correctlist)
        kaipucorrect(correctids)


def send_nosee():
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f" 检测到开普云有新类型错敏，机器人无法修改，请人工核实。\n<@WuXiaoLong>\n"
        }
    }

    key = conf.key_cs
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))

    # 输出响应结果
    print(response.text)


def send_correct_msg(correctlist):
    key = conf.key_cs

    wb = Workbook()

    ws = wb.active

    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格设置标题和表头
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '机器人改错记录  ' + endDate_str
    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font

    ws.append(['错敏词', '修改词', '快照', '地址', '父地址', '错敏类型', '父标题'])
    # 设置新添加的行的字体为加粗
    for row in [ws[row + 1]]:
        for cell in row:
            cell.font = conf.header_font

    # 遍历字典

    for tup in correctlist:
        ws.append(tup)
    ws.column_dimensions['A'].width = 20
    # 设置列A的宽度为50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    # 保存文件

    formatted_date = endDate.strftime('%Y-%m-%d %H_%M_%S')  # 使用下划线代替冒号
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 指定一个子目录名，例如 'saved_files'
    subdirectory = 'coreectlist'

    # 创建子目录的完整路径
    subdirectory_path = os.path.join(script_dir, subdirectory)

    # 如果子目录不存在，则创建它
    if not os.path.exists(subdirectory_path):
        os.makedirs(subdirectory_path)

    # 构建最终的文件路径
    filename = os.path.join(subdirectory_path, conf.correct_name + formatted_date + '.xlsx')
    wb.save(filename)
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key

    # 加载现有的Excel文件
    wb = load_workbook(filename)
    # 获取活动工作表

    # 将工作簿保存到一个字节流中
    output = BytesIO()
    wb.save(output)
    base_filename = os.path.basename(filename)

    # 准备发送文件，确保只使用文件名
    files = {"file": (base_filename, output.getvalue())}
    # 准备发送文件

    response = requests.post(
        'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=' + key + '&type=file',
        files=files)

    print(response.text)

    # 检查是否有'media_id'在响应中
    if 'media_id' in response.json():
        # 发送消息
        headers = {"Content-Type": "application/json"}
        message = {
            "msgtype": "file",
            "file": {
                "media_id": response.json()["media_id"]
            }
        }
        response = requests.post(url, headers=headers, json=message)

        # 输出响应结果
        print(response.text)
    else:
        print("No media_id in response")


if __name__ == '__main__':
    print(token)
    dealcuo()
    # print(token)
    # formatted_date = endDate.strftime('%Y-%m-%d %H_%M_%S')  # 使用下划线代替冒号
    # filename = conf.correct_name + formatted_date + '.xlsx'
    # print(filename)
    # getjsseion.hello()
    # hudongcontent.search_message_by_id('20240620115221974',1,bz_gov_id,jid)
    # dealtext.save_data()

