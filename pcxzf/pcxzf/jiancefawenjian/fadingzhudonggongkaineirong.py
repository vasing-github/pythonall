# 引入 Beautiful Soup 模块
from bs4 import BeautifulSoup
import requests
import json
import datetime
import conf
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
from openpyxl import Workbook
from threed_record import qm


def getAllbumen(prox):
    url = 'http://www.scpc.gov.cn/public/column/6601841?type=4&action=rel'
    res = conf.make_request_get(url)
    # 提取所有 ul 标签下的所有 li 元素
    li_elements = []
    for ul in res[1].find_all('ul', class_='clearfix'):
        li_elements += ul.find_all('li')

    # 将 li 元素转换为字典
    bumenList = {}
    for li in li_elements:
        a_element = li.find('a')
        if a_element is not None:
            title = a_element.get('title')
            href = a_element.get('href')
            bumenList[title] = href

    # 输出字典
    print(bumenList)
    return bumenList


def getMenniu(prox, bumen):
    res = conf.make_request_get(bumen)
    # 提取 id 为 fdzd_gknr 的 div 标签下的所有 a 标签
    a_elements = res[1].find('div', {'id': 'fdzd_gknr'}).find_all('a')
    # 输出所有 a 标签的文本内容和链接地址
    menuList = {}
    for a in a_elements:
        # print(f'{a.text}: {a["href"]}')
        if a is not None:
            title = a.text
            href = a["href"]
            menuList[title] = href
    return menuList


def getTime(prox, menu):
    res = conf.make_request_get(menu)
    li_elements = res[1].find_all('li', {'class': 'rq'})
    # print(li_elements)
    if len(li_elements) == 1:
        return '无'
    second_li_element = li_elements[1]

    return second_li_element.text


def gettittle(href):
    res = conf.make_request_get(href)
    li_elements = res[1].find_all('li', {'class': 'mc'})
    # print(li_elements)
    if len(li_elements) <= 1:
        return '无'

    return li_elements[1].find('a').text


def cacul(href):
    # 这里传进来的其实是字符串2023-11-7
    date = datetime.datetime.strptime(href, '%Y-%m-%d').date()
    delta = datetime.date.today() - date
    return delta.days


def makeMsg(ws, msg):
    # 获取当前的最大行数，并加1
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格设置标题和表头
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '一、法定主动公开内容栏目监测'
    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font
    ws.append(['1.栏目超期情况监测:以下位置栏目已超期，请责任部门更新'])
    ws.append(['单位名称', '栏目名称', '更新标题', '最后更新时间', '超期天数'])
    # 设置新添加的行的字体为加粗
    for row in [ws[row + 1], ws[row + 2]]:
        for cell in row:
            cell.font = conf.header_font

    # 遍历字典
    for key, value in msg.items():
        for menu, tup in value.items():
            # 计算天数
            days = cacul(tup[0])
            # 将数据添加到工作表
            ws.append([key, menu, tup[1], tup[0], days])


def make_no_content_list(ws, no_content_list):
    # 填充该栏目监测的其他问题
    ws.append([''])
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 1

    ws.append(['2.死链监测：以下单位对应栏目无内容，请上传内容或审核发布'])
    ws.append(['单位名称', '栏目名称'])
    for r in [row, row + 1]:
        for cell in ws[r]:
            cell.font = conf.header_font

    for dup in no_content_list:
        ws.append([dup[0], dup[1]])


def make_tj_miss_year(ws, tj_miss_year_dic):
    # 填充该栏目监测的其他问题
    ws.append([''])
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 1

    ws.append(['3.统计信息缺失年份：以下单位缺失对应年份统计信息，请补充'])
    ws.append(['单位名称', '缺失年份'])
    for r in [row, row + 1]:
        for cell in ws[r]:
            cell.font = conf.header_font

    for key, value in tj_miss_year_dic.items():
        # 将key和value中的元素合并成一个列表
        row_data = [key] + value
        # 添加到工作表中
        ws.append(row_data)


def makeExcel(msg, no_content_list, tj_miss_year_dic):
    # 加载现有的Excel文件
    wb = load_workbook(conf.xlsx_name)
    # 获取活动工作表
    ws = wb.active

    makeMsg(ws, msg)
    make_no_content_list(ws, no_content_list)
    make_tj_miss_year(ws, tj_miss_year_dic)

    # 保存文件
    wb.save(conf.xlsx_name)


# 这个方法不是直接取网页数据了，是取网页中的js方法，以及参数，用它来请求每一页的数据，如果js为空  说明就只有一页数据，那就拿到这一页的日期返回去，如果不为空，就用gettimebypage方法返回所有日期列表
def getkeywork(prox, menu, time_list):
    res = conf.make_request_get(menu)
    # 查找所有的<script>标签
    scripts = res[1].find_all('script')
    # 遍历所有的<script>标签
    for script in scripts:
        # 检查<script>标签的内容是否为None
        if script.string is not None:
            # 检查<script>标签的内容是否包含感兴趣的JavaScript代码
            if 'Ls.pagination' in script.string:
                # 打印<script>标签的内容

                # 使用正则表达式提取url和data
                url_match = re.search(r'url: "(.*?)"', script.string)
                data_match = re.search(r'data : ({.*?})', script.string, re.DOTALL)
                page_count_match = re.search(r'pageCount:(\d+)', script.string)

                if url_match and data_match and page_count_match:
                    # 提取url
                    url = url_match.group(1)

                    # 提取data
                    data_str = data_match.group(1)

                    page_count = page_count_match.group(1)

                    return getTimeListByPage(time_list, url, data_str, page_count)

    li_elements = res[1].find_all('li', {'class': 'rq'})
    if len(li_elements) == 1:
        return '无'
    time_list.extend(li_elements)
    return time_list


def getTimeListByPage(time_list, url, data_str, pagecount):
    # 存储所有页面的数据
    all_pages_data = []
    # 使用正则表达式提取siteId，catId和file
    site_id_match = re.search(r'siteId : "(.*?)"', data_str)
    cat_id_match = re.search(r'catId: "(.*?)"', data_str)
    file_match = re.search(r'file:"(.*?)"', data_str)
    organ_id_match = re.search(r'organId: "(.*?)"', data_str)

    if site_id_match and cat_id_match and file_match and organ_id_match:
        # 提取siteId，catId和file
        site_id = site_id_match.group(1)
        cat_id = cat_id_match.group(1)
        file = file_match.group(1)
        organid = organ_id_match.group(1)
        # print(f"siteId: {site_id}, catId: {cat_id}, file: {file}")
        for i in range(1, int(pagecount) + 1):
            # 请求的URL
            if not url.startswith('http'):
                url = 'http://www.scpc.gov.cn' + url
            # print(url)
            # 请求的参数
            params = {
                "labelName": "publicInfoList",
                "siteId": site_id,
                "pageSize": "18",
                "pageIndex": i,
                "action": "list",
                "isDate": "true",
                "dateFormat": "yyyy-MM-dd",
                "length": "50",
                "organId": organid,
                "isDetail": "",
                "type": "4",
                "catId": cat_id,
                "cId": "",
                "result": "暂无相关信息",
                "file": file
            }

            # 发送请求
            response = requests.get(url, params=params)

            # 检查响应状态码
            if response.status_code == 200:
                # 将页面数据添加到列表中
                all_pages_data.append(response.text)
                soup = BeautifulSoup(response.text, 'html.parser')
                li_elements = soup.find_all('li', {'class': 'rq'})

                time_list.extend(li_elements)
            else:
                print(f"请求失败，状态码：{response.status_code}")

    return time_list


def get_all_tj_time(href, time_list, key, menu, no_content_list):
    upTime = getkeywork(None, href, time_list)
    # print('uptime', upTime)
    # print(len(time_list))
    if upTime == '无':
        no_content_list.append((key, menu))
    return upTime


def deal_tj_year(final_dic):
    new_data = {}

    # 遍历旧字典
    for unit, dates in final_dic.items():
        if unit == '平昌县人民政府办公室':
            continue
        # 初始化年份集合
        years = set()
        # 遍历日期
        for date in dates:
            # 提取年份
            year = date.text.split('-')[0]
            # 添加年份到集合
            years.add(year)

        # 根据单位设置检查的年份范围
        if unit == '四川平昌经济开发区管理委员会' or unit == '平昌县金宝街道办事处' or unit == '平昌县江家口镇人民政府':
            check_years = range(2020, 2023)

        elif unit == '平昌县商务局' or unit == '平昌县退役军人事务局' or unit == '平昌县信访局' or unit == '平昌县医疗保障局':
            check_years = range(2019, 2023)
        else:
            check_years = range(2018, 2023)

        # 检查是否包含指定年份
        missing_years = [str(year) for year in check_years if str(year) not in years]
        # # 检查是否包含2018-2022年
        # missing_years = [str(year) for year in range(2018, 2023) if str(year) not in years]
        # 如果有缺少的年份，添加到新字典
        if missing_years:
            new_data[unit] = missing_years

    # 打印新字典

    return new_data


def deal_menu_dic(menu_Dic, key, menu, href, no_content_list, final_dic):
    upTime = getTime(None, href)

    if upTime == '无':
        no_content_list.append((key, menu))
        return
    days = cacul(upTime)

    # 判断时间差是否超过配置表中规定超期时间
    if days >= conf.up_time_conf[menu]:
        title = gettittle(href)
        menu_Dic[menu] = upTime, title
    if menu_Dic:
        final_dic[key] = menu_Dic


def startMain():
    bumen_Mennu_list = conf.bumen_Mennu_list
    final_dic = {}
    no_content_list = []
    tj_final_dic = {}

    for key, value in bumen_Mennu_list.items():
        print(key)
        menu_Dic = {}
        tj_time_list = []
        for menu, href in value.items():
            print(menu)
            if key == '平昌县人民政府办公室':
                continue

            if menu == '统计信息':
                # 收集统计信息的所有发文时间
                tj_final_dic[key] = get_all_tj_time(href, tj_time_list, key, menu, no_content_list)

            deal_menu_dic(menu_Dic, key, menu, href, no_content_list, final_dic)

    tj_miss_year_dic = deal_tj_year(tj_final_dic)

    # final_dic = {'平昌县经济和信息化局': {'统计信息': ('2023-09-21', '2023年1-8月全县规上工业实现产值22.54亿元                                            '), '重大民生信息': ('2023-09-21', '全县新建5G基站146座                                            '), '其他法定信息': ('2023-04-20', '《四川省中小企业特色产业集群培育认定管理暂行办法》政策解读                                            ')}, '平昌县财政局': {'其他法定信息': ('2023-09-12', '平昌县财政局机关生活垃圾分类实施方案                                            ')}, '平昌县住房和城乡建设局': {'统计信息': ('2023-10-12', '2023年1-9月资金争取情况                                            '), '重大项目': ('2023-10-12', '2023年项目储备情况                                            ')}, '平昌县退役军人事务局': {'统计信息': ('2023-08-28', '2023年4月—6月退役军人建档立卡情况统计                                            '), '重大民生信息': ('2023-05-26', '平昌县退役军人及其他优抚对象统计情况                                            '), '其他法定信息': ('2023-07-14', '拥军优抚政策解读                                            ')}, '平昌县林业局': {'统计信息': ('2023-09-25', '平昌县林业发展情况统计                                            '), '重大项目': ('2023-09-25', '平昌县储备林建设项目正式启动                                            ')}, '平昌县医疗保障局': {'统计信息': ('2023-09-22', '关于2023年第八批居民、职工门诊特殊疾病认定结果的通知                                            '), '行政许可/处罚': ('2023-09-28', '关于同意将平昌县友邦药业有限公司金太阳光店等52家连锁定点零售药店纳入门诊统筹管理的公告                                            '), '其他法定信息': ('2023-09-21', '平昌县医疗保障局关于将巴中市普济医药连锁有限公司平昌三五三分店等16家诊所药店纳入医保协议定点管理的公示                                            ')}, '平昌县金宝街道办事处': {'其他法定信息': ('2023-09-22', '金宝街道烟花爆竹零售经营门店（2024-2025）布点规划方案                                            ')}, '平昌县岳家镇人民政府': {'统计信息': ('2023-10-07', '岳家镇新增入市项目储备库2个                                            ')}, '平昌县龙岗镇人民政府': {'统计信息': ('2023-08-29', '龙岗镇二季度蔬菜及特种作物生产统计表                                            ')}, '平昌县佛楼镇人民政府': {'其他法定信息': ('2023-09-22', '佛楼镇便民服务中心运行管理办法                                            ')}, '平昌县响滩镇人民政府': {'统计信息': ('2023-08-24', '响滩镇2023年大春经济作物种植面积统计表                                            '), '重大项目': ('2023-08-24', '响滩镇肉牛产业项目建设有序推进                                            ')}, '平昌县大寨镇人民政府': {'政府工作报告': ('2023-01-03', '大寨镇人民政府2023年政府工作报告                                            '), '统计信息': ('2023-08-24', '平昌县大寨镇医保缴费工作推进情况                                            '), '预算/决算': ('2023-01-12', '大寨镇2022年财政预算执行情况及2023年财政预算                                            '), '重大项目': ('2023-08-24', '平昌县大寨镇福申社区污水处理工程项目建设情况                                            ')}, '平昌县青云镇人民政府': {'统计信息': ('2023-10-10', '青云镇辣椒种植统计                                            '), '重大项目': ('2023-10-10', '金宝街道至青云道路改建工程道路建设                                            '), '重大民生信息': ('2023-10-10', '关于做好2024年度城乡居民基本医疗保险参保工作的通知                                            '), '其他法定信息': ('2023-10-10', '农村公益性岗位补贴申领办事指南                                            ')}, '平昌县澌岸镇人民政府': {'规划计划': ('2023-01-11', '2023年工作计划                                            ')}, '平昌县云台镇人民政府': {'统计信息': ('2023-08-24', '云台镇2023年5-8月安全工作开展情况统计                                            ')}, '平昌县岩口镇人民政府': {'重大项目': ('2023-07-18', '岩口镇2023年1—6月集体经济问效报表                                            ')}, '平昌县镇龙镇人民政府': {'统计信息': ('2023-09-25', '镇龙镇2023年上半年主要蔬菜作物播种面积统计                                            '), '重大项目': ('2023-09-25', '红旗水库除险加固工程基本竣工                                            ')}, '平昌县望京镇人民政府': {'统计信息': ('2023-07-25', '望京镇关于成立望京镇第五次全国经济普查领导小组的通知                                            '), '重大民生信息': ('2023-08-25', '望京镇2023年燃气煤气消防安全应急预案                                            ')}, '平昌县江家口镇人民政府': {'统计信息': ('2023-09-21', '江家口镇2023年度大春经济作物产量统计                                            '), '其他法定信息': ('2023-09-21', '江家口镇2023年春季雨露计划补助名单                                            ')}}
    #
    # tj_miss_year_dic = {'平昌县农业农村局': ['2020'], '平昌县商务局': ['2018', '2019'], '平昌县退役军人事务局': ['2018', '2019'], '平昌县市场监督管理局': ['2018', '2019', '2020'], '平昌县乡村振兴局': ['2020'], '平昌县信访局': ['2018', '2019', '2020'], '平昌县医疗保障局': ['2018', '2019', '2021', '2022'], '四川平昌经济开发区管理委员会': ['2018', '2019'], '平昌县金宝新区管理委员会': ['2018', '2019', '2020', '2021'], '平昌县金宝街道办事处': ['2018', '2019'], '平昌县龙岗镇人民政府': ['2019', '2020'], '平昌县佛楼镇人民政府': ['2021'], '平昌县响滩镇人民政府': ['2021'], '平昌县大寨镇人民政府': ['2018'], '平昌县驷马镇人民政府': ['2018', '2019', '2020', '2021'], '平昌县粉壁镇人民政府': ['2019', '2020'], '平昌县元山镇人民政府': ['2020'], '平昌县云台镇人民政府': ['2020', '2021', '2022'], '平昌县邱家镇人民政府': ['2019', '2021'], '平昌县泥龙镇人民政府': ['2018'], '平昌县岩口镇人民政府': ['2018', '2019', '2020'], '平昌县镇龙镇人民政府': ['2018', '2019', '2020']}
    # no_content_list = [('平昌县交通运输局', '机关简介'), ('平昌县岳家镇人民政府', '机关简介')]
    # print('-----------------------------')
    # print(final_dic)
    # print('-----------------------------')
    # print(no_content_list)
    # print('-----------------------------')
    # print(tj_miss_year_dic)
    # print('-----------------------------')

    makeExcel(final_dic, no_content_list, tj_miss_year_dic)
    qm.add_item({conf.menu_over_update: final_dic})
    qm.add_item({conf.miss_tj_year: tj_miss_year_dic})


if __name__ == '__main__':

    # getAllbumen(None)
    bumenlist = {'平昌县人民政府办公室': '/public/column/6601841?type=2', '平昌县发展和改革局': '/public/column/6601861?type=2', '平昌县经济和信息化局': '/public/column/6601881?type=2', '平昌县教育科技局': '/public/column/6601901?type=2', '平昌县公安局': '/public/column/6601921?type=2', '平昌县民政局': '/public/column/6601941?type=2', '平昌县司法局': '/public/column/6601961?type=2', '平昌县财政局': '/public/column/6601981?type=2', '平昌县人力资源和社会保障局': '/public/column/6602001?type=2', '平昌县自然资源和规划局': '/public/column/6602021?type=2', '平昌县住房和城乡建设局': '/public/column/6602041?type=2', '平昌县交通运输局': '/public/column/6602061?type=2', '平昌县水利局': '/public/column/6602081?type=2', '平昌县农业农村局': '/public/column/6602101?type=2', '平昌县商务局': '/public/column/6602121?type=2', '平昌县文化广播电视体育和旅游局': '/public/column/6602141?type=2', '平昌县卫生健康局': '/public/column/6602161?type=2', '平昌县退役军人事务局': '/public/column/6602181?type=2', '平昌县应急管理局': '/public/column/6602201?type=2', '平昌县审计局': '/public/column/6602221?type=2', '平昌县市场监督管理局': '/public/column/6602241?type=2', '巴中市平昌生态环境局': '/public/column/6602261?type=2', '平昌县统计局': '/public/column/6602281?type=2', '平昌县信访局': '/public/column/6602321?type=2', '平昌县林业局': '/public/column/6602341?type=2', '平昌县医疗保障局': '/public/column/6602361?type=2', '平昌县行政审批和数据局': '/public/column/6602381?type=2', '平昌县综合行政执法局': '/public/column/6602401?type=2', '四川平昌经济开发区管理委员会': '/public/column/6604364?type=2', '平昌县金宝新区管理委员会': '/public/column/6602841?type=2', '平昌县佛头山管理委员会': '/public/column/6602881?type=2', '平昌县同州街道办事处': '/public/column/6603021?type=2', '平昌县江口街道办事处': '/public/column/6603041?type=2', '平昌县金宝街道办事处': '/public/column/6604081?type=2', '平昌县白衣镇人民政府': '/public/column/6603081?type=2', '平昌县涵水镇人民政府': '/public/column/6603541?type=2', '平昌县岳家镇人民政府': '/public/column/6603561?type=2', '平昌县西兴镇人民政府': '/public/column/6603521?type=2', '平昌县龙岗镇人民政府': '/public/column/6603241?type=2', '平昌县土垭镇人民政府': '/public/column/6603461?type=2', '平昌县佛楼镇人民政府': '/public/column/6603581?type=2', '平昌县响滩镇人民政府': '/public/column/6603101?type=2', '平昌县大寨镇人民政府': '/public/column/6603501?type=2', '平昌县驷马镇人民政府': '/public/column/6603061?type=2', '平昌县青云镇人民政府': '/public/column/6603621?type=2', '平昌县兰草镇人民政府': '/public/column/6603221?type=2', '平昌县澌岸镇人民政府': '/public/column/6603361?type=2', '平昌县粉壁镇人民政府': '/public/column/6603321?type=2', '平昌县得胜镇人民政府': '/public/column/6603261?type=2', '平昌县元山镇人民政府': '/public/column/6603161?type=2', '平昌县灵山镇人民政府': '/public/column/6603741?type=2', '平昌县土兴镇人民政府': '/public/column/6603481?type=2', '平昌县云台镇人民政府': '/public/column/6603181?type=2', '平昌县三十二梁镇人民政府': '/public/column/6604061?type=2', '平昌县板庙镇人民政府': '/public/column/6603201?type=2', '平昌县邱家镇人民政府': '/public/column/6603301?type=2', '平昌县笔山镇人民政府': '/public/column/6603121?type=2', '平昌县泥龙镇人民政府': '/public/column/6603661?type=2', '平昌县岩口镇人民政府': '/public/column/6603701?type=2', '平昌县镇龙镇人民政府': '/public/column/6603141?type=2', '平昌县望京镇人民政府': '/public/column/6603281?type=2', '平昌县江家口镇人民政府': '/public/column/6604071?type=2'}
    bumen_menu_list = {}
    for key, value in bumenlist.items():
        menu_list = getMenniu(None,value)
        bumen_menu_list[key] = menu_list

    print(bumen_menu_list)

    # startMain()
    # bumen_Mennu_list = conf.bumen_Mennu_list
    # final_dic = {}
    # no_content_list = []
    # tj_final_dic = {}
    #
    # for key, value in bumen_Mennu_list.items():
    #     print(key)
    #     menu_Dic = {}
    #     tj_time_list = []
    #     for menu, href in value.items():
    #         print(menu)
    #         if key == '平昌县人民政府办公室':
    #             continue
    #         if menu == '统计信息':
    #             # 收集统计信息的所有发文时间
    #             tj_final_dic[key] = get_all_tj_time(href, tj_time_list, key, menu, no_content_list)
    #
    #         deal_menu_dic(menu_Dic, key, menu, href, no_content_list, final_dic)
    #
    # tj_miss_year_dic = deal_tj_year(tj_final_dic)
    # print(tj_miss_year_dic)
