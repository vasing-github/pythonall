# 引入 Beautiful Soup 模块
from bs4 import BeautifulSoup
import requests
import json
import datetime
import conf
from openpyxl import load_workbook
from threed_record import qm
from openpyxl.styles import Font
# 这个py是用来检测政府信息公开工作年报的，62个部门都有公开
# 这个文件不能直接运行第三步，要运行第二步找到所有单位所有公开的年份，第三步只是分析年份是否有缺失，以及对应年份链接是否正确，发布时间是否正确

def getAllbumen(prox):
    url = 'http://www.scpc.gov.cn/public/column/6601841?type=4&action=rel'
    res = conf.make_request_get(url)

    # 提取所有 ul 标签下的所有 li 元素
    li_elements = []
    for ul in res[1].find_all('ul', class_='clearfix'):
        li_elements += ul.find_all('li')

    # 将 li 元素转换为字典
    bumenList = {}
    for li in li_elements:
        a_element = li.find('a')
        if a_element is not None:
            title = a_element.get('title')
            href = a_element.get('href')
            bumenList[title] = href

    return bumenList
def getMenniu(prox,bumen):
    res = conf.make_request_get(bumen)
    # 提取 id 为 fdzd_gknr 的 div 标签下的所有 a 标签
    a_elements = res[1].find('ul', {'id': 'timeFrame'}).find_all('a')
    # 输出所有 a 标签的文本内容和链接地址

    menuList={}
    for a in a_elements:
        # print(f'{a.text}: {a["href"]}')
        if a is not None:
            title = a.text
            href = a["href"]
            menuList[title] = href
    return menuList


def cacul(href):
    # 这里传进来的其实是字符串2023-11-7
    date = datetime.datetime.strptime(href, '%Y-%m-%d').date()
    delta = datetime.date.today() - date
    return delta.days


def startMain():
    bumenList = conf.bumenList
    #
    # # 这是第二步 通过单位找政府信息公开年报所有列表得到的是部门名称下所有年份对应的地址
    bumen_Mennu_list = {}
    for key, value in bumenList.items():
        menuDic = getMenniu(None, value)
        bumen_Mennu_list[key] = menuDic
    # print(bumen_Mennu_list)


    missing_years_dict = {}
    for unit, years in bumen_Mennu_list.items():
        if unit == '四川平昌经济开发区管理委员会' or unit == '平昌县金宝街道办事处' or unit == '平昌县江家口镇人民政府':
            years_to_check = [str(year) + '年' for year in range(2020, 2023)]
        elif unit == '平昌县商务局':
            years_to_check = [str(year) + '年' for year in range(2019, 2023)]
        else:
            years_to_check = [str(year) + '年' for year in range(2018, 2023)]

        missing_years = [year for year in years_to_check if year not in years.keys()]
        if missing_years:
            # print(f"单位 {unit} 缺少以下年份: {', '.join(missing_years)}")
            missing_years_dict[unit] = missing_years

    print(missing_years_dict)
    # missing_years_dict = {'平昌县商务局': ['2018年', '2019年'], '平昌县退役军人事务局': ['2018年', '2019年'], '平昌县信访局': ['2018年', '2019年'], '平昌县医疗保障局': ['2018年', '2019年'], '平昌县金宝街道办事处': ['2018年', '2019年'], '平昌县驷马镇人民政府': ['2020年'], '平昌县江家口镇人民政府': ['2018年', '2019年']}
    qm.add_item({conf.zf_year_report: missing_years_dict})
    makeExcel(missing_years_dict)


def makeExcel(msg):
    # 加载现有的Excel文件
    wb = load_workbook(conf.xlsx_name)
    # 获取活动工作表
    ws = wb.active

    # 获取当前的最大行数，并加2
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '二、政府信息公开年报栏目缺失情况监测'

    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font
    ws.append(['单位名称','缺失年份'])
    # 设置新添加的行的字体为加粗
    for cell in ws[row+1]:
        cell.font = conf.header_font
    # 遍历字典
    for key, value in msg.items():
        # 将key和value中的元素合并成一个列表
        row_data = [key] + value
        # 添加到工作表中
        ws.append(row_data)

    # 保存文件
    wb.save(conf.xlsx_name)

if __name__ == '__main__':

    startMain()

