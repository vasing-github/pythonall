from datetime import datetime
from io import BytesIO

import conf
import fadingzhudonggongkaineirong
import gongkainianbao
import jigouzhineng
import requests
from jiancefawenjian import zhengce, jicengzwgk
from openpyxl import Workbook
from openpyxl import load_workbook
from threed_record import qm


def get_current_time():
    # 获取当前时间
    now = datetime.now()
    # 格式化时间为指定格式
    formatted_time = now.strftime("%Y-%m-%d %H:%M:%S")
    return formatted_time


def sendMsg():
    key = conf.key_choose
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key

    # 上传文件
    filename = conf.xlsx_name

    # 加载现有的Excel文件
    wb = load_workbook(filename)
    # 获取活动工作表

    # 将工作簿保存到一个字节流中
    output = BytesIO()
    wb.save(output)

    # 准备发送文件
    files = {"file": (filename, output.getvalue())}
    response = requests.post(
        'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=' + key + '&type=file',
        files=files)

    print(response.text)

    # 检查是否有'media_id'在响应中
    if 'media_id' in response.json():
        # 发送消息
        headers = {"Content-Type": "application/json"}
        message = {
            "msgtype": "file",
            "file": {
                "media_id": response.json()["media_id"]
            }
        }
        response = requests.post(url, headers=headers, json=message)

        # 输出响应结果
        print(response.text)
    else:
        print("No media_id in response")


def end_excel():
    wb = load_workbook(conf.xlsx_name)
    ws = wb.active
    ws['j1'] = '检测时间：'+get_current_time()
    # 在h3列写入"栏目更新检测标准"
    ws['j3'] = '栏目更新检测标准'
    ws['j3'].font = conf.header_font
    # 在F4写入"栏目名称"，在G4写入"超期时间"
    ws['j4'] = '栏目名称'
    ws['k4'] = '超期时间'

    # 从F5开始遍历字典up_time_conf并填充内容
    start_row = 5
    for i, (menu, time) in enumerate(conf.up_time_conf.items(), start=start_row):
        ws['j{}'.format(i)] = menu
        ws['k{}'.format(i)] = time

        # 设置列A的宽度为50
    ws.column_dimensions['A'].width = 40
    # 设置列A的宽度为50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.column_dimensions['j'].width = 20
    ws.column_dimensions['k'].width = 20

    ws['j49'] = '基层政务公开栏目检测标准:'
    ws['j49'].font = conf.header_font
    ws['j50'] = '1.每个事项的总内容不少于4条'
    ws['j51'] = '2.最新更新时间不大于90天'
    ws['j52'] = '3.每相邻两条信息之间更新时间不大于90天'
    ws['j53'] = '注意：除办事服务等不易发布的内容外，其余链接形式的内容应尽量粘贴复制而不是直接跳转，否则机器人无法监测'

    # 保存更改
    wb.save(conf.xlsx_name)


def make_new_excel():
    # 创建一个新的工作簿
    wb = Workbook()
    # 保存文件
    wb.save(conf.xlsx_name)


def start_main():
    # 填充法定主动公开内容超期情况
    fadingzhudonggongkaineirong.startMain()
    gongkainianbao.startMain()
    jigouzhineng.startMain()
    zhengce.startMain()
    # jicengzwgk.startMain()

    # 给表格增加一些注释
    end_excel()


if __name__ == '__main__':
    make_new_excel()
    start_main()
    qm.add_item({conf.finish_score: None})
    sendMsg()
    qm.stop()
