import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import conf
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
from openpyxl import Workbook
import time
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Cookie': 'search_history=%E5%8E%95%E6%89%80%E9%9D%A9%E5%91%BD; SHIROJSESSIONID=8b77c272-388b-411e-8dc2-f5115ea2bd40',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'
}


def makeExcel(len_less_four_list, uptime_more_ninteen, each_more_ninteen):
    # 加载现有的Excel文件
    wb = load_workbook(conf.xlsx_name)

    # 获取活动工作表
    ws = wb.active

    # 获取当前的最大行数，并加1
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格设置标题和表头
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1).value = '四、基层政务公开专栏监测'

    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font
    ws.append(['1.以下位置内容条数少于4条'])
    # 设置新添加的行的字体为加粗
    for cell in ws[row + 1]:
        cell.font = conf.header_font
    ws.append(['单位名称', '领域', '一级事项', '二级事项'])
    for item in len_less_four_list:
        content = [conf.jczwgk_area_comp_dic[item[0]]] + list(item)
        ws.append(content)

    row = ws.max_row
    ws.cell(row=row + 2, column=1).value = '2.以下位置最近更新时间大于90天'
    # ws.append(['2.以下位置最近更新时间大于90天'])
    for cell in ws[row + 2]:
        cell.font = conf.header_font
    ws.append(['单位名称', '领域', '一级事项', '二级事项'])
    for item in uptime_more_ninteen:
        content = [conf.jczwgk_area_comp_dic[item[0]]] + list(item)
        ws.append(content)

    row = ws.max_row
    ws.cell(row=row + 2, column=1).value = '3.以下位置相邻文章发布时间间隔大于90天'
    # ws.append(['3.以下位置相邻文章发布时间间隔大于90天'])
    for cell in ws[row + 2]:
        cell.font = conf.header_font
    ws.append(['单位名称', '领域', '一级事项'])
    for item in each_more_ninteen:
        # for menu in item[0]:
        for info in item[1]:
            row_data = list(item[0]) + [info]
            # 使用 append() 方法将新列表添加到工作表中
            content = [conf.jczwgk_area_comp_dic[item[0][0]]] + row_data
            ws.append(content)
            # ws.append(row_data)

    # 保存文件
    wb.save(conf.xlsx_name)


def get_all_type_menu():
    url = 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ggwhfw/index.html'

    res = conf.make_request_get(url)
    div_elment = res[1].find('div', {'class': 'classify_name ly_box'})
    # print(div_elment)
    a_element = div_elment.find_all('a')
    area_type_dic = {}
    if a_element is not None:
        for _a in a_element:
            title = _a.get('title')
            href = _a.get('href')
            area_type_dic[title] = href
    return area_type_dic


def get_first_munu(url):
    res = conf.make_request_get(url)
    div_elment = res[1].find('div', {'class': 'classify_name sx_box'})
    a_element = div_elment.find_all('a')
    first_menu_dic = {}
    if a_element is not None:
        for _a in a_element:
            title = _a.get('title')
            href = _a.get('href')
            first_menu_dic[title] = href
    return first_menu_dic


def get_second_menu(url):
    res = conf.make_request_get(url)
    # 提取curFirstLevelChildId的值
    cur_first_level_child_id = res[1].find(id='curFirstLevelChildId')
    # 如果能拿到这个值说明一级事项给的链接没有跳转，如果拿不到说明没有二级事项，一级事项就跳转了
    if cur_first_level_child_id is not None:
        cur_first_level_child_id = res[1].find(id='curFirstLevelChildId')['value']
        url = "http://www.scpc.gov.cn/site/label/8888"
        params = {
            "isJson": "true",
            "id": cur_first_level_child_id,
            "isChild": "true",
            "labelName": "columnTree"
        }
        # response = requests.get(url, params=params)
        res = conf.make_request_get(url, params)
        second_menu_dic = {}
        if res[0].text != 'null':
            for item in res[0].json():
                name = item.get('name')
                html_path = item.get('htmlPath')
                if not name or not html_path:
                    continue  # 如果name或html_path为空，则跳过当前循环
                second_menu_dic[name] = html_path
        else:
            # 返回1表示没有二级事项且不跳转
            return 1, cur_first_level_child_id
    else:
        # 返回2表示没有二级事项，但是跳转了
        return 2, 0
    return 3, second_menu_dic


def get_page(url):
    # 正则表达式
    pattern = r'www\.scpc\.gov\.cn(.*)'
    # 搜索匹配的路径
    match = re.search(pattern, url)
    if match:
        url = match.group(1)
    res = conf.make_request_get(url)
    # 二级事项跳转目录不尽相同，如果div有listnews，直接在这个页面找，并且有些会翻页,第一步找含有listnews的能否翻页，第二部找是否有listnews，第三步找没有listnews的
    # 找到包含pageCount的<script>标签
    script_tag = res[1].find('script', string=re.compile('pageCount'))
    # 使用正则表达式从<script>标签的内容中提取pageCount的值
    if script_tag is not None:
        page_count_match = re.search(r'pageCount\s*:\s*(\d+)', script_tag.string)
        page_count = page_count_match.group(1)
        script_tag = res[1].find('script', string=re.compile(r"http://www.scpc.gov.cn/content/column/\d+\?pageIndex="))

        # 使用正则表达式从<script>标签的内容中提取URL的值
        if script_tag is not None:
            url_match = re.search(r"(http://www.scpc.gov.cn/content/column/\d+\?pageIndex=)", script_tag.string)
            if url_match:
                url = url_match.group(1)
                print(f"提取的URL是: {url}")
                return deal_page(page_count, url)
    else:
        # 如果没有pagecount，可能是只有一页，所以传值1，但也可能是不跳转，所以没有pagecount，这个时候在基层政务公开这个页面找右侧列表
        div_elment = res[1].find('div', {'class': 'navjz clearfix'})
        if div_elment is not None:
            a_element = div_elment.find_all('a')
            page_news_dic = {}
            if a_element is not None:
                for _a in a_element:
                    title = _a.get('title')
                    href = _a.get('href')
                    page_news_dic[title] = href
                return deal_area_page_news(page_news_dic)
        else:
            return deal_page(1, url)


def deal_area_page_news(page_news_dic):
    list_all_page = []
    for title, url in page_news_dic.items():
        res = conf.make_request_get(url)
        div_elment = res[1].find('div', {'class': 'newsinfo clearfix'})
        if div_elment is None:
            continue
        date_span = div_elment.find('span', class_='sp')
        # 提取日期字符串
        date_str = date_span.text.split('：')[1].strip()
        # 只获取日期部分，不包括时间
        date_only = date_str.split(' ')[0]
        list_all_page.append((date_only, title))
    return list_all_page


def deal_page(page_count, url):
    list_all_page = []
    for i in range(1, int(page_count) + 1):
        new_url = url + '?pageIndex=' + str(i)
        # print(new_url)
        res = conf.make_request_get(new_url)
        # 如果这里的网页请求不到，说明就是在基层政务公开这个专栏发的，没有跳转到其他地方链接，这个时候要用网页里面获取页面内容的js方法请求
        # 有些页面加了pageindex仍然能访问，但实际却并不会跳转，这一部分再上面处理
        if res[0].status_code != 404:

            div_elment = res[1].find('div', {'class': 'listnews'})
            if div_elment is not None:
                li_elment = div_elment.find_all('li')
                for li in li_elment:

                    if li.find('span', class_='right date') is not None:
                        date = li.find('span', class_='right date').text
                        # 提取标题
                        title = li.find('a', class_='left')['title']
                        list_all_page.append((date, title))
        else:
            deal_has_sencond_not_redict(url, list_all_page)

    return list_all_page


def deal_has_sencond_not_redict(url, list_all_page):
    res = conf.make_request_get(url)
    div_elment = res[1].find('div', {'class': 'navjz clearfix'})
    if div_elment is not None:
        a_element = div_elment.find_all('a')
        page_news_dic = {}
        if a_element is not None:
            for _a in a_element:
                title = _a.get('title')
                href = _a.get('href')
                page_news_dic[title] = href
        return deal_area_page_news(page_news_dic)


def deal_first_menu_not_direct(url, cur_id):
    url = "http://www.scpc.gov.cn/site/label/8888"
    # 请求的数据
    data = {
        "id": cur_id,
        "isChild": "true",
        "siteId": 6787191,
        "pageSize": 100,
        "pageIndex": 1,
        "length": 54,
        "isDate": "false",
        "isHit": "false",
        "isLine": "true",
        "lineCount": 5,
        "dateFormat": "yyyy-MM-dd",
        "file": "/c3/jh/articleNewsPageList_new_jczw",
        "labelName": "articleNewsPageList"
    }

    # 发送请求
    response = requests.post(url, data=data)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'html.parser')

    a_element = soup.find_all('a')
    page_news_dic = {}
    if a_element is not None:
        for _a in a_element:
            title = _a.get('title')
            href = _a.get('href')
            # 如果标题已经在字典中，就添加到对应的列表
            if title in page_news_dic:
                timestamp = int(time.time() * 1e6)  # 获取当前时间戳，精度为微秒
                title = f"{title}_{timestamp}"
            page_news_dic[title] = href
    return deal_area_page_news(page_news_dic)


def deal_first_menu_direct(url):
    res = conf.make_request_get(url)
    script_tag = res[1].find('script', string=re.compile('pageCount'))
    # 使用正则表达式从<script>标签的内容中提取pageCount的值
    if script_tag is not None:
        page_count_match = re.search(r'pageCount\s*:\s*(\d+)', script_tag.string)
        page_count = page_count_match.group(1)
        script_tag = res[1].find('script', string=re.compile(r"http://www.scpc.gov.cn/content/column/\d+\?pageIndex="))

        # 使用正则表达式从<script>标签的内容中提取URL的值
        if script_tag is not None:
            url_match = re.search(r"(http://www.scpc.gov.cn/content/column/\d+\?pageIndex=)", script_tag.string)
            if url_match:
                url = url_match.group(1)
                print(f"提取的URL是: {url}")
                return deal_page(page_count, url)
    else:
        # 如果没有pagecount，可能是只有一页，所以传值1，但也可能是不跳转，所以没有pagecount，这个时候在基层政务公开这个页面找右侧列表
        div_elment = res[1].find('div', {'class': 'navjz clearfix'})
        if div_elment is not None:
            a_element = div_elment.find_all('a')
            page_news_dic = {}
            if a_element is not None:
                for _a in a_element:
                    title = _a.get('title')
                    href = _a.get('href')
                    page_news_dic[title] = href
                return deal_area_page_news(page_news_dic)
        else:
            return deal_page(1, url)


def judgement_news(articles):
    is_less_than_four = len(articles) < 4

    # 判断最新更新时间是否大于90天
    latest_update_date = datetime.strptime(articles[0][0], '%Y-%m-%d') if articles else None
    is_latest_update_more_than_ninety = latest_update_date and (
                datetime.now() - latest_update_date > timedelta(days=90))

    # 判断相邻两个文章的更新时间是否大于90天
    list_each_more_ninety = [
        f"文章 '{articles[i][1]}' 和 '{articles[i + 1][1]}' 之间的更新时间大于90天"
        for i in range(len(articles) - 1)
        if
        datetime.strptime(articles[i][0], '%Y-%m-%d') - datetime.strptime(articles[i + 1][0], '%Y-%m-%d') > timedelta(
            days=90)
    ]

    return is_less_than_four, is_latest_update_more_than_ninety, list_each_more_ninety


def deal_information(menu_dup, error_news_dup, len_less_four_list, uptime_more_ninteen, each_more_ninteen):
    if error_news_dup[0]:
        len_less_four_list.append(menu_dup)
    if error_news_dup[1]:
        uptime_more_ninteen.append(menu_dup)
    if len(error_news_dup[2]) > 0:
        each_more_ninteen.append((menu_dup, error_news_dup[2]))


def startMain():
    len_less_four_list = []
    uptime_more_ninteen = []
    each_more_ninteen = []
    # area_type_dic = get_all_type_menu()
    area_type_dic = {'交通运输': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/jtys/index.html'}
    # print(area_type_dic)
    # area_type_dic = {'公共文化服务': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ggwhfw/index.html', '就业领域': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/jyly/index.html', '涉农补贴': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/snbt/index.html', '食品药品': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/spyp/index.html', '社会救助': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/shjz/index.html', '卫生健康': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/wsjk/index.html', '养老服务': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ylfw/index.html', '义务教育': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ywjy/index.html', '公共法律': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ggfl/index.html', '税收管理': 'https://sichuan.chinatax.gov.cn/col/col12057/index.html?number=A0015', '广播电视和网络视听': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/gbdshwlst/index.html', '旅游领域': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ly/index.html', '社会保险': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/shbx/index.html', '自然资源': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/zrzy/index.html', '城市综合执法': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/cszhzf/index.html', '户籍管理': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/hjgl/index.html', '财政预决算': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/czyjs/index.html', '市政服务': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/szfw/index.html', '农村危房改造': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ncwfgz/index.html', '国有土地上房屋征收与补偿': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/gytdsfwzsybc/index.html', '保障性住房': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/bzxzf/index.html', '新闻出版版权': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/xwcbbq/index.html', '生态环境': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/sthj/index.html', '统计领域': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/tjly/index.html', '公共资源交易': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/ggzyjy/index.html', '交通运输': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/jtys/index.html', '扶贫领域': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/fply/index.html', '重大建设项目': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/zdjsxm/index.html', '安全生产': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/aqsc/index.html', '救灾领域': 'http://www.scpc.gov.cn/ztzl/jczwgk/gkly/jzly/index.html'}

    for area, href in area_type_dic.items():
        print('领域：', area)
        if area == '税收管理':
            continue
        first_menu_dic = get_first_munu(href)
        for first_key, first_href in first_menu_dic.items():
            print('一级事项：', first_key)
            second_menu_dic = get_second_menu(first_href)
            print(second_menu_dic)
            if second_menu_dic[0] == 3:
                for second_key, second_href in second_menu_dic[1].items():
                    print('二级事项：', second_key)
                    list_all_news = get_page(second_href)
                    print(list_all_news)
                    error_news_dup = judgement_news(list_all_news)
                    deal_information((area, first_key, second_key), error_news_dup, len_less_four_list,
                                     uptime_more_ninteen, each_more_ninteen)
            # 没有二级事项
            elif second_menu_dic[0] == 1:
                # 返回1表示没有二级事项且不跳转
                list_all_news = deal_first_menu_not_direct(first_href, second_menu_dic[1])

                error_news_dup = judgement_news(list_all_news)
                deal_information((area, first_key), error_news_dup, len_less_four_list, uptime_more_ninteen,
                                 each_more_ninteen)

            elif second_menu_dic[0] == 2:
                # 返回2表示没有二级事项还跳转了
                list_all_news = deal_first_menu_direct(first_href)

                error_news_dup = judgement_news(list_all_news)
                deal_information((area, first_key), error_news_dup, len_less_four_list, uptime_more_ninteen,
                                 each_more_ninteen)

    makeExcel(len_less_four_list, uptime_more_ninteen, each_more_ninteen)


if __name__ == '__main__':
    startMain()