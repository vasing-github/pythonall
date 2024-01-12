from bs4 import BeautifulSoup
import requests
import json
import datetime
import conf
from openpyxl import load_workbook
from openpyxl.styles import Font
# /**
# 这是从机构职能页进入检查机关简介
# **/

def getAllbumen(prox,jigou_dic):
    url = 'http://www.scpc.gov.cn/jgzn/index.html'
    data = {}
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Cookie':'search_history=%E5%8E%95%E6%89%80%E9%9D%A9%E5%91%BD; SHIROJSESSIONID=8b77c272-388b-411e-8dc2-f5115ea2bd40',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'
    }
    response = requests.get(url, headers=headers, data=data,proxies=prox)

    response.encoding = 'utf-8'
    # 以 Beautiful Soup 解析 HTML 程序码
    soup = BeautifulSoup(response.text, 'html.parser')
    for all_a in soup.find_all('a', {'class': 'left'}):
        title = all_a.get('title')
        href = all_a.get('href')
        if title in jigou_dic:
            sendMsgError('f ## 机构职能页死链监测：\n\n 以下机构重复，检测是否多发布一条',f'f {title}--{href}')
            continue
        jigou_dic[title] = href

def getzhishu(prox,jigou_dic):
    url = 'http://www.scpc.gov.cn/ljglzy/jgznljgl/xzfgzbmgwh/6790771.js?num=99&ect=1697104935796'

    data = {}
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    response = requests.get(url, headers=headers, data=data,proxies=prox)
    response.encoding = 'utf-8'
    # 以 Beautiful Soup 解析 HTML 程序码
    soup = BeautifulSoup(response.text, 'html.parser')
    for all_a in soup.find_all('a'):
        title = all_a.get('title')
        href = all_a.get('href')
        if title in jigou_dic:
            sendMsgError(f'## 机构职能页死链监测：\n\n 以下机构重复，检测是否多发布一条',f' {title}--{href}')
            continue
        jigou_dic[title] = href


def sendMsgError(tittle,content):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f"{tittle}\n\n\n{content}"
        }
    }
    # 定义企业微信机器人的webhook地址
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=425a53f6-696e-46c1-9d4a-fae8940b136f"

    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))
    # 输出响应结果
    print(response.text)
def sendMsg(tittle,content):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f"{tittle}\n\n\n{content}"
        }
    }
    # 定义企业微信机器人的webhook地址  测试群
    # url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=425a53f6-696e-46c1-9d4a-fae8940b136f"
    # 正式群
    url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=6c00ba33-68ab-403e-bab2-2b9134a7d7f6'

    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))
    # 输出响应结果
    print(response.text)

def modifgyHref(key,url, not_open_list ,error_state_list,zzgk_error_list ):

    if not url.startswith('http'):
        url = 'http://www.scpc.gov.cn' + url
    data = {}
    headers = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    response = requests.get(url, headers=headers, data=data)
    response.encoding = 'utf-8'

    if '文章id' in response.text and '未发布' in response.text:
        not_open_list.append(key)
        # sendMsgError(f'## 机构职能页监测：\n\n以下位置未发布',f'{key}--{response.text}')
        # return 1
    if response.status_code != 200:
        error_state_list.append(key)
        # sendMsgError(f'## 机构职能页监测：\n\n以下位置链接有误，请检查',f'{key}--{response.status_code}')
        # return 1
    soup = BeautifulSoup(response.text, 'html.parser')
    div_elment = soup.find_all('div', {'class': 'xxgk-wzcon'})

    for element in div_elment:
        if '政务公开' not in element.get_text():
            zzgk_error_list.append(key)

    return not_open_list,error_state_list,zzgk_error_list
def sendMsgErrorList(titlle,list):
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": f"{titlle}\n\n"
        }
    }
    for key in list:
        data["markdown"]["content"] += f"### --{key}\n\n"

    # 定义企业微信机器人的webhook地址

    # url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=425a53f6-696e-46c1-9d4a-fae8940b136f"
        # 正式群
    url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=6c00ba33-68ab-403e-bab2-2b9134a7d7f6'


    # 发送HTTP POST请求
    response = requests.post(url, data=json.dumps(data))
    # 输出响应结果
    print(response.text)

def startMain():
    jigou_dic = {}
    getAllbumen(None, jigou_dic)
    getzhishu(None, jigou_dic)
    # print(len(jigou_dic))
    if len(jigou_dic) != 74:
        sendMsg(f'## 机构职能页监测：\n\n', f'机构数量不等于74.检查是否多发少发')
    not_open_list = []
    error_state_list = []
    zzgk_error_list = []
    for key, value in jigou_dic.items():
        modifgyHref(key, value, not_open_list, error_state_list, zzgk_error_list)

    # not_open_list= ['镇龙镇', ' 县卫生健康局']
    # error_state_list=['白衣镇']
    # zzgk_error_list = ['青云镇','江口水乡管理局']

    makeExcel(not_open_list,error_state_list,zzgk_error_list)

def makeExcel(not_open_list,error_state_list,zzgk_error_list):
    # 加载现有的Excel文件
    wb = load_workbook(conf.xlsx_name)

    # 获取活动工作表
    ws = wb.active

    # 获取当前的最大行数，并加1
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格设置标题和表头
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '三、机构职能页监测'

    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font
    ws.append(['1.以下位置未发布，请管理员处理'])
    # if not_open_list:
    ws.append(not_open_list)
    ws.append(['2.以下位置链接状态错误，请管理员检查'])
    # if not_open_list:
    ws.append(error_state_list)
    ws.append(['3.以下单位机关简介中无领导分管政务公开，请对应部门修改'])
    # if not_open_list:
    ws.append(zzgk_error_list)

    # 保存文件
    wb.save(conf.xlsx_name)


if __name__ == '__main__':
    startMain()



