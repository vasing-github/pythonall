# 引入 Beautiful Soup 模块
import datetime

import conf
from openpyxl import load_workbook
from threed_record import qm

# 这个py是用来检测政府信息政策栏目的，抄的公开年报的py

def getAllbumen(prox):
    url = 'http://www.scpc.gov.cn/public/column/6601841?type=4&action=rel'
    res = conf.make_request_get(url)

    # 提取所有 ul 标签下的所有 li 元素
    li_elements = []
    for ul in res[1].find_all('ul', class_='clearfix'):
        li_elements += ul.find_all('li')

    # 将 li 元素转换为字典
    bumenList = {}
    for li in li_elements:
        a_element = li.find('a')
        if a_element is not None:
            title = a_element.get('title')
            href = a_element.get('href')
            bumenList[title] = href

    return bumenList


def getMenniu(prox, bumen):
    res = conf.make_request_get(bumen)
    # 提取 id 为 fdzd_gknr 的 div 标签下的所有 a 标签
    a_tag = res[1].find('a', id='policyLink')
    # 输出所有 a 标签的文本内容和链接地址
    if a_tag:
        href = a_tag.get('href')
        print(f'The href attribute is: {href}')
    else:
        print('No element with id="policyLink" found.')

    return href


def cacul(href):
    # 这里传进来的其实是字符串2023-11-7
    date = datetime.datetime.strptime(href, '%Y-%m-%d').date()
    delta = datetime.date.today() - date
    return delta.days


def getTime(prox, menu):
    res = conf.make_request_get(menu)
    li_elements = res[1].find_all('li', {'class': 'rq'})
    # print(li_elements)
    if len(li_elements) == 1:
        return '无'
    second_li_element = li_elements[1]

    return second_li_element.text


def gettittle(href):
    res = conf.make_request_get(href)
    li_elements = res[1].find_all('li', {'class': 'mc'})
    # print(li_elements)
    if len(li_elements) <= 1:
        return '无'

    return li_elements[1].find('a').text


def startMain():
    # bumenList = conf.bumenList
    # bumen_zhengce_list = {}
    # for key, value in bumenList.items():
    #     if key == '平昌县人民政府办公室':
    #         continue
    #     menuDic = getMenniu(None, value)
    #     bumen_zhengce_list[key] = menuDic
    # print(bumen_zhengce_list)

    # bumen_zhengce_list = {
    #
    #     '平昌县金宝新区管理委员会': '/public/column/6602841?type=4&action=list&nav=2&sub=0&catId=6717326',
    # }

    bumen_zhengce_list = conf.bumen_zhengce_list
    zc_over_list = []
    for key, href in bumen_zhengce_list.items():
        upTime = getTime(None, href)
        days = cacul(upTime)

        # 判断时间差是否超过配置表中规定超期时间
        if days >= conf.up_time_conf['政策']:
            title = gettittle(href)
            print(title)
            zc_over_list.append((key, title, upTime, days))

    qm.add_item({conf.zc_over_time: zc_over_list})
    print(zc_over_list)


    # zc_over_list = [('平昌县商务局', '平昌县财政局平昌县商务局关于印发《平昌县2021年国家级电子商务进农村综合示范建设项目专项资金使用管理办法》的通知                                            ', '2023-03-07', 437), ('平昌县审计局', '四川省审计整改结果公告办法                                            ', '2023-04-10', 403), ('平昌县驷马镇人民政府', '驷马镇安全生产大提升专项行动工作方案                                            ', '2023-04-06', 407), ('平昌县得胜镇人民政府', '平昌县得胜镇人民政府关于印发《得胜镇丰收村2023年乡村振兴重点帮扶村工作方案》的通知                                            ', '2023-05-12', 371)]
    makeExcel(zc_over_list)


def makeExcel(msg):
    # 加载现有的Excel文件
    wb = load_workbook(conf.xlsx_name)
    # 获取活动工作表
    ws = wb.active

    # 获取当前的最大行数，并加2
    row = ws.max_row if ws.max_row == 1 else ws.max_row + 2

    # 合并单元格
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = '四、政府信息政策栏超期情况'

    # 设置合并后单元格的字体样式
    ws.cell(row=row, column=1).font = conf.tittle_font
    ws.append(['单位名称', '更新标题', '最后更新时间', '超期天数'])
    # 设置新添加的行的字体为加粗
    for cell in ws[row + 1]:
        cell.font = conf.header_font
    # 遍历字典
    for dup in msg:

        ws.append([dup[0],dup[1],dup[2],dup[3]])

    # 保存文件
    wb.save(conf.xlsx_name)


if __name__ == '__main__':
    startMain()
