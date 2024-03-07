import queue
import threading
import conf
import pymysql
import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO

class QueueManager:
    def __init__(self):
        self.q = queue.Queue()
        self.thread = threading.Thread(target=self.listen_queue)
        self.thread.daemon = True
        self.thread.start()

        self.switch = {
            conf.menu_over_update: self.process_menu_over,
            conf.miss_tj_year: self.process_tj_miss,
            conf.zf_year_report: self.process_zf_year_miss,
            conf.finish_score: self.process_finish_score,
        }
        # 创建数据库连接和游标
        self.connection = pymysql.connect(**conf.database)
        self.cursor = self.connection.cursor()

    def listen_queue(self):
        while True:
            item = self.q.get()
            if item is None:
                break
            if conf.is_record:
                break
            self.process_item(item)
            self.q.task_done()

    def add_item(self, item):
        self.q.put(item)

    def execute_query(self, sql, params=None):
        # print(sql)
        # 执行查询语句
        self.cursor.execute(sql, params)
        # 获取查询结果
        result = self.cursor.fetchall()
        return result

    def update_data(self, sql, params=None):
        # 执行更新语句
        self.cursor.execute(sql, params)
        # 提交更新
        self.connection.commit()

    def insert_data(self, sql, up_record_list):
        print(sql)
        self.cursor.executemany(sql, up_record_list)
        # 提交更新
        self.connection.commit()

    def process_menu_over(self, val):
        up_record_list = []
        up_assessment_record = []
        for danwei, over in val.items():

            for men, dup in over.items():
                sql_search = f"SELECT * FROM demerit_record WHERE company_name = '{danwei}' AND reason = '{conf.menu_over_update}' AND path = '{men}'"
                reslt = self.execute_query(sql_search)

                # print(reslt)
                if reslt:
                    # 如果记录表存在就要添加扣分信息
                    up_assessment_record.append((danwei, conf.menu_over_update, men, -1))
                up_record_list.append((danwei, conf.menu_over_update, men))
        print(up_record_list)
        # 删除记录表
        sql_up_demerit_record = f'DELETE FROM demerit_record WHERE reason = "{conf.menu_over_update}"'
        self.update_data(sql_up_demerit_record)
        # 插入记录表新值
        sql_insert_record = "INSERT INTO demerit_record (company_name, date,reason, path) VALUES (%s, CURDATE(),%s, %s)"
        self.insert_data(sql_insert_record, up_record_list)
        # 插入扣分表
        sql_insert_assessment_record = "INSERT INTO assessment_record (company_name, date,score_change_reason, path,score) VALUES (%s, CURDATE(),%s, %s,  %s)"
        self.insert_data(sql_insert_assessment_record, up_assessment_record)

    def process_list_miss(self, val, reason):
        up_record_list = []
        up_assessment_record = []
        for danwei, l in val.items():

            for year in l:
                sql_search = f"SELECT * FROM demerit_record WHERE company_name = '{danwei}' AND reason = '{reason}' AND path = '{year}'"
                reslt = self.execute_query(sql_search)

                # print(reslt)
                if reslt:
                    # 如果记录表存在就要添加扣分信息
                    up_assessment_record.append((danwei, reason, year, -1))
                up_record_list.append((danwei, reason, year))
        print(up_record_list)
        # 删除记录表
        sql_up_demerit_record = f'DELETE FROM demerit_record WHERE reason = "{reason}"'
        self.update_data(sql_up_demerit_record)
        # 插入记录表新值
        sql_insert_record = "INSERT INTO demerit_record (company_name, date,reason, path) VALUES (%s, CURDATE(),%s, %s)"
        self.insert_data(sql_insert_record, up_record_list)
        # 插入扣分表
        sql_insert_assessment_record = "INSERT INTO assessment_record (company_name, date,score_change_reason, path,score) VALUES (%s, CURDATE(),%s, %s,  %s)"
        self.insert_data(sql_insert_assessment_record, up_assessment_record)

    def process_zf_year_miss(self, val):
        self.process_list_miss(val, conf.zf_year_report)

    def process_tj_miss(self, val):
        self.process_list_miss(val, conf.miss_tj_year)

    def process_finish_score(self, val):
        sql = """
        SELECT company_name, SUM(score) AS total_score
        FROM assessment_record
        WHERE date = CURDATE()
        GROUP BY company_name;
        """
        res = self.execute_query(sql)

        sql = """
                SELECT *
                FROM assessment_record
                WHERE date = CURDATE();
            """
        res2 = self.execute_query(sql)
        self.sendMsg(res,res2)

    def process_item(self, item):
        for key, value in item.items():
            func = self.switch.get(key)
            if func:
                func(value)

    def sendMsg(self, res,res2):

        # 创建一个新的工作簿
        wb = Workbook()
        # 获取活动工作表
        ws = wb.active
        row = ws.max_row if ws.max_row == 1 else ws.max_row + 2
        # 合并单元格设置标题和表头
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).value = '一、本周各单位分数汇总'
        # 设置合并后单元格的字体样式
        ws.cell(row=row, column=1).font = conf.tittle_font
        ws.append(['单位名称', '本周扣分合计'])
        # 设置新添加的行的字体为加粗
        for row in ws[row + 1]:
            row.font = conf.header_font

        # 对结果按照扣分进行排序
        sorted_res = sorted(res, key=lambda row: row[1])

        # 遍历排序后的结果
        for row in sorted_res:
            ws.append([row[0], row[1]])

        row = ws.max_row if ws.max_row == 1 else ws.max_row + 2
        # 合并单元格设置标题和表头
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).value = '二、本周各单位扣分详情'
        # 设置合并后单元格的字体样式
        ws.cell(row=row, column=1).font = conf.tittle_font
        ws.append(['单位名称', '扣分时间', '分值', '扣分原因', '路径', '备注1', '备注2'])
        # 设置新添加的行的字体为加粗
        for row in ws[row + 1]:
            row.font = conf.header_font
        for row in res2:
            ws.append([row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        ws.column_dimensions['A'].width = 40
        # 设置列A的宽度为50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20

        # 保存文件
        wb.save(conf.xlsx_score_name)
        self.send_excel()

    def send_excel(self):
        key = conf.record_key_choose
        url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=" + key
        # 上传文件
        filename = conf.xlsx_score_name
        wb = load_workbook(filename)
        # 获取活动工作表

        # 将工作簿保存到一个字节流中
        output = BytesIO()
        wb.save(output)

        # 准备发送文件
        files = {"file": (filename, output.getvalue())}
        response = requests.post(
            'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=' + key + '&type=file',
            files=files)

        print(response.text)

        # 检查是否有'media_id'在响应中
        if 'media_id' in response.json():
            # 发送消息
            headers = {"Content-Type": "application/json"}
            message = {
                "msgtype": "file",
                "file": {
                    "media_id": response.json()["media_id"]
                }
            }
            response = requests.post(url, headers=headers, json=message)

            # 输出响应结果
            print(response.text)
        else:
            print("No media_id in response")
    def stop(self):
        # 添加一个None到队列来通知线程退出
        self.q.put(None)
        self.thread.join()
        self.cursor.close()
        self.connection.close()


qm = QueueManager()
